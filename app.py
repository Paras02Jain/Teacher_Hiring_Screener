"""
app.py — EduHire Teacher Resume Screener v8
============================================
New in v8:
  - Candidate Pipeline Stage Tracker (Applied → Screened → Interview Scheduled → Feedback Pending → Offer/Reject)
  - Stall Detection — flags candidates with no activity for X days
  - Cold Candidate Monitor — alerts before candidates go silent
  - Call/Meeting Notes Logger — timestamped notes per candidate
  - Gmail Notifications — stage-change emails sent via Gmail SMTP to hiring manager

Previous (v7):
  - Auto Email Draft Generator — personalised shortlist & rejection emails
  - Groq AI email generation with smart rule-based fallback
  - Bulk generate + download all emails as .txt
  - School name & HR name configurable from sidebar
"""

import streamlit as st
import re, io
from backend import ResumeScreener, detect_jd_bias
from chatbot import get_chatbot_response, get_suggested_questions

# ── New modules (v9 additions) ────────────────────────────────────────────────
try:
    from fairness_audit import run_fairness_audit
    _FAIRNESS_AVAILABLE = True
except ImportError:
    _FAIRNESS_AVAILABLE = False

try:
    from executive_pdf import generate_executive_pdf, reportlab_available
    _PDF_AVAILABLE = True
except ImportError:
    _PDF_AVAILABLE = False

try:
    from jd_ab_tester import run_ab_test
    _AB_AVAILABLE = True
except ImportError:
    _AB_AVAILABLE = False
from email_generator import generate_email_draft, generate_all_emails, send_email
from pipeline_tracker import (
    PIPELINE_STAGES, STAGE_COLORS, STAGE_ICONS,
    init_pipeline, get_stage, set_stage, add_note,
    get_stalled_candidates, get_cold_candidates,
    get_pipeline_summary, days_since, _cand_key,
)

try:
    from calendar_scheduler import CalendarScheduler, generate_slots, _GOOGLE_AVAILABLE, _duration_label
    _CALENDAR_MODULE = True
except ImportError:
    _CALENDAR_MODULE = False
    _GOOGLE_AVAILABLE = False

try:
    from rag_engine import build_index
    _RAG_AVAILABLE = True
except ImportError:
    _RAG_AVAILABLE = False

try:
    from interview_brief import generate_brief, enrich_brief_with_ai, brief_to_text, _fmt_ts as _brief_fmt_ts
    _BRIEF_MODULE = True
except ImportError:
    _BRIEF_MODULE = False

from visualizations import (
    chart_score_bar, chart_score_vs_experience,
    chart_skills_frequency, chart_keyword_gap,
    chart_credentials_donut, chart_radar_top3,
    chart_experience_distribution,
    chart_score_distribution, chart_candidate_heatmap,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="EduHire · Teacher Resume Screener",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Theme — rich teal/indigo with warm cream canvas ─────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:wght@300;400;500;600;700&display=swap');

/* Main canvas */
.stApp,[data-testid="stAppViewContainer"],[data-testid="stMain"],
.main,.block-container{
  background: linear-gradient(135deg,#0D1B2A 0%,#1B2838 40%,#0F3460 100%) !important;
  color:#E8EFF8 !important;
  font-family:'DM Sans',sans-serif !important;
}
p,span,div,label,h1,h2,h3,h4,li{color:#E8EFF8 !important;}
.stMarkdown p,.stMarkdown span,.stMarkdown div{color:#E8EFF8 !important;}
.stSlider label p,.stCheckbox label p{color:#D1DCF0 !important;}

/* Textarea */
textarea{
  background-color:#12253A !important;color:#E8EFF8 !important;
  border:1.5px solid #2A4A6B !important;border-radius:10px !important;
}

/* File uploader */
[data-testid="stFileUploader"]{
  background: linear-gradient(135deg,#0D2137,#122D4A) !important;
  border:2px dashed #3A9BD5 !important;border-radius:12px !important;
}
[data-testid="stFileUploader"] *{color:#D1DCF0 !important;}

/* Buttons */
.stButton>button{
  background: linear-gradient(135deg,#0EA5E9,#0F766E) !important;
  color:#FFFFFF !important;border:none !important;border-radius:12px !important;
  font-weight:700 !important;font-size:1rem !important;padding:14px 40px !important;
  font-family:'DM Sans',sans-serif !important;
}
.stButton>button:hover{
  background: linear-gradient(135deg,#38BDF8,#14B8A6) !important;
  transform:translateY(-1px);box-shadow:0 8px 24px rgba(14,165,233,.35) !important;
}
[data-testid="stDownloadButton"] button{
  background: linear-gradient(135deg,#10B981,#059669) !important;
  color:#FFFFFF !important;border:none !important;border-radius:10px !important;
  font-weight:700 !important;padding:10px 28px !important;
}

/* Sidebar */
[data-testid="stSidebar"]{
  background: linear-gradient(180deg,#0A1628 0%,#0D1F3C 50%,#0A1628 100%) !important;
  border-right:1px solid #1E3A5F !important;
}
[data-testid="stSidebar"] *{color:#A8C5E8 !important;}
[data-testid="stSidebar"] input{
  background-color:#0F2540 !important;color:#E8EFF8 !important;
  border:1px solid #2A4A6B !important;border-radius:8px !important;
}
[data-testid="stSidebar"] .stMarkdown p{color:#A8C5E8 !important;}

/* Expanders */
[data-testid="stExpander"]{
  background: rgba(13,33,64,0.7) !important;
  border:1px solid #1E3A5F !important;border-radius:12px !important;
}
[data-testid="stExpander"] summary p{color:#A8C5E8 !important;}

/* Alerts */
.stSuccess,.stWarning,.stError,.stInfo{border-radius:10px !important;}

/* Remove default Streamlit top padding */
.block-container{padding-top:2rem !important;}
</style>
""", unsafe_allow_html=True)

# ─── Colour palette ───────────────────────────────────────────────────────────
C = {
    "ink":       "#0D1B2A",  "white":    "#FFFFFF",  "cream":    "#E8EFF8",
    "gold":      "#F59E0B",  "gold_lt":  "#FEF3C7",  "gold_bg":  "#78350F",
    "sage":      "#10B981",  "sage_lt":  "#D1FAE5",  "sage_dk":  "#065F46",
    "rose":      "#F43F5E",  "rose_lt":  "#FFE4E6",
    "amber":     "#F59E0B",  "amber_lt": "#FEF3C7",
    "silver":    "#94A3B8",  "bronze":   "#D97706",
    "border":    "#1E3A5F",  "gray":     "#64748B",
    "sky":       "#0EA5E9",  "sky_lt":   "#E0F2FE",
    "teal":      "#0F766E",  "teal_lt":  "#CCFBF1",
    "navy":      "#0D1B2A",  "card":     "rgba(13,33,64,0.85)",
    "card2":     "rgba(15,52,96,0.6)",
}

def _clean_text(t):
    """Strip ALL HTML tags and neutralise remaining angle brackets so
    the text can safely be embedded inside an unsafe_allow_html block."""
    t = str(t)
    prev = None
    while prev != t:
        prev = t
        t = re.sub(r'<[^>]*>', '', t)
    t = (t.replace("&lt;", "<").replace("&gt;", ">").replace("&amp;", "&")
          .replace("&quot;", '"').replace("&#39;", "'").replace("&nbsp;", " "))
    prev = None
    while prev != t:
        prev = t
        t = re.sub(r'<[^>]*>', '', t)
    t = t.replace("<", "&lt;").replace(">", "&gt;")
    return re.sub(r'\s+', ' ', t).strip()

def tag(label, bg, fg):
    return (f'<span style="display:inline-block;background:{bg};color:{fg};'
            f'font-size:.72rem;font-weight:700;padding:3px 10px;'
            f'border-radius:100px;margin:2px 3px 2px 0;">{label}</span>')

def green_tags(items):
    return "".join(tag(t, "#0F3D2A", "#6EE7B7") for t in items)

def rose_tags(items):
    return "".join(tag(t, "#3D0A14", "#FDA4AF") for t in items)

def kw_tags(items, matched=True):
    bg = "#0A2E1F" if matched else "#2E0A12"
    fg = "#6EE7B7" if matched else "#FDA4AF"
    icon = "✓ " if matched else "✗ "
    return "".join(tag(icon + t, bg, fg) for t in items)

def score_pill(score, bg, fg):
    return (f'<span style="display:inline-flex;align-items:center;gap:4px;'
            f'background:{bg};color:{fg};font-weight:800;font-size:.88rem;'
            f'padding:5px 16px;border-radius:100px;border:1.5px solid {fg}55;">⬡ {score:.1f}% match</span>')

def score_bar(score, bar_color):
    pct = min(score, 100)
    return (f'<div style="background:rgba(255,255,255,.1);border-radius:100px;height:7px;'
            f'width:150px;margin-top:7px;overflow:hidden;">'
            f'<div style="background:{bar_color};height:7px;border-radius:100px;'
            f'width:{pct:.0f}%;"></div></div>')

def method_badge(method):
    if method == "sentence_transformer":
        return tag("🧠 sentence-transformer", "#1E3A5F", "#7DD3FC")
    elif method == "tfidf":
        return tag("📊 TF-IDF", "#3B2F0A", "#FCD34D")
    return tag("🔤 keyword", "#3B2F0A", "#FCD34D")

def premier_badge(is_premier):
    if is_premier:
        return tag("🏛️ Premier Institution", "#2D1A5E", "#C4B5FD")
    return tag("Institution", "#1E2A3D", "#64748B")


def render_ats_breakdown(ats: dict) -> str:
    """Render a compact ATS score breakdown bar chart in HTML."""
    if not ats:
        return ""
    categories = [
        ("🎓 Education",      "education",     "#0EA5E9"),
        ("💼 Experience",     "experience",    "#10B981"),
        ("🛠️ Skills",         "skills",        "#F59E0B"),
        ("🔑 Keywords",       "keywords",      "#8B5CF6"),
        ("📜 Certification",  "certification", "#EC4899"),
    ]
    weighted = ats.get("weighted_ats", 0)
    overall  = ats.get("overall_score", 0)

    bars = ""
    for label, key, color in categories:
        cat   = ats.get(key, {})
        score = cat.get("score", 0)
        weight= int(cat.get("weight", 0) * 100)
        bars += (
            f'<div style="margin-bottom:9px;">' +
            f'<div style="display:flex;justify-content:space-between;margin-bottom:3px;">' +
            f'<span style="font-size:.72rem;color:#A8C5E8;font-weight:600;">{label}</span>' +
            f'<span style="font-size:.72rem;color:{color};font-weight:700;">{score:.0f}% <span style="color:#64748B;font-weight:400;">({weight}% weight)</span></span>' +
            f'</div>' +
            f'<div style="background:rgba(255,255,255,.08);border-radius:100px;height:6px;">' +
            f'<div style="background:{color};height:6px;border-radius:100px;width:{min(score,100):.0f}%;' +
            f'transition:width .4s ease;"></div></div>' +
            f'</div>'
        )

    ats_color = "#10B981" if weighted >= 60 else "#F59E0B" if weighted >= 35 else "#F43F5E"
    html = (
        f'<div style="background:rgba(8,20,40,0.7);border:1px solid rgba(30,58,95,.6);' +
        f'border-radius:12px;padding:16px 20px;margin-top:14px;">' +
        f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;">' +
        f'<span style="font-size:.82rem;font-weight:700;color:#A8C5E8;">📊 ATS Score Breakdown</span>' +
        f'<span style="font-size:.88rem;font-weight:800;color:{ats_color};">ATS: {weighted:.0f}%</span>' +
        f'</div>' +
        bars +
        f'</div>'
    )
    return html

def divider():
    return (f'<hr style="border:none;height:1px;margin:32px 0;'
            f'background:linear-gradient(90deg,transparent,{C["sky"]},transparent);">')


# ─── Hero ─────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style="background:linear-gradient(135deg,#0A1628 0%,#0D2137 50%,#0F3460 100%);
     border-radius:20px;padding:52px 56px;margin-bottom:32px;position:relative;overflow:hidden;">
  <div style="position:absolute;inset:0;background:repeating-linear-gradient(
    45deg,transparent,transparent 40px,rgba(14,165,233,.04) 40px,rgba(14,165,233,.04) 41px);"></div>
  <div style="position:absolute;right:48px;bottom:10px;font-size:6rem;opacity:.07;">🎓</div>
  <div style="position:absolute;top:-60px;right:-60px;width:220px;height:220px;
    background:radial-gradient(circle,rgba(14,165,233,.15),transparent 70%);border-radius:50%;"></div>
  <div style="position:relative;">
    <span style="display:inline-flex;align-items:center;background:rgba(14,165,233,.15);
      border:1px solid rgba(14,165,233,.35);color:{C["sky"]};font-size:.72rem;font-weight:700;
      padding:4px 14px;border-radius:100px;letter-spacing:1.5px;text-transform:uppercase;
      margin-bottom:16px;">🎓 AI-Powered · Local Scoring + HF AI Reasoning</span>
    <h1 style="font-size:3rem;font-weight:900;color:#FFF;margin:0 0 10px;letter-spacing:-1px;
      line-height:1.1;font-family:'DM Serif Display',Georgia,serif;">
      EduHire <span style="color:{C["sky"]};">Screener</span>
    </h1>
    <p style="font-size:1rem;color:rgba(168,197,232,.8);font-weight:300;margin:0;">
      Local sentence-transformer scoring &nbsp;·&nbsp;
      AI chatbot via Groq (Free Llama/Mixtral) &nbsp;·&nbsp;
      Premier institution detection &nbsp;·&nbsp;
      Full Excel export
    </p>
  </div>
</div>
""", unsafe_allow_html=True)

# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    groq_key = st.text_input(
        "🔑 Groq API Key", type="password",
        placeholder="gsk_...",
        help="Free API key from console.groq.com — powers Llama 3, Mixtral, Gemma. No credit card required."
    )
    groq_model = st.selectbox(
        "🤖 AI Model",
        [
            "llama-3.3-70b-versatile",
            "llama-3.1-8b-instant",
            "mixtral-8x7b-32768",
            "gemma2-9b-it",
        ],
        index=0,
        format_func=lambda m: {
            "llama-3.3-70b-versatile": "⭐ Llama 3.3 70B (Best quality)",
            "llama-3.1-8b-instant":    "⚡ Llama 3.1 8B (Fastest)",
            "mixtral-8x7b-32768":      "🧠 Mixtral 8x7B (Great reasoning)",
            "gemma2-9b-it":            "💎 Gemma 2 9B (Google)",
        }.get(m, m),
        help="All models are FREE on Groq. Llama 3.3 70B gives the best answers."
    )
    use_ai_just = st.checkbox("🤖 Use AI Justification", value=True,
                              help="If unchecked, uses rule-based justification. Faster but less detailed.")
    st.markdown("---")
    st.markdown("**🧠 Scoring Method**")
    st.markdown("Score computed **locally** via `sentence-transformers/all-MiniLM-L6-v2`.")
    st.markdown("---")
    st.markdown("**💬 AI Chatbot**")
    st.markdown("Powered by **Groq + Llama/Mixtral** — ask *anything*.")
    if not groq_key:
        st.info("Add your free Groq key above to unlock full AI chat.")
    st.markdown("---")
    st.markdown("**📧 Email Generator**")
    school_name = st.text_input(
        "🏫 School Name",
        value="Our School",
        help="Used in generated email drafts as sender's organisation."
    )
    hr_name = st.text_input(
        "👤 HR / Sender Name",
        value="The Hiring Team",
        help="Sign-off name in generated emails."
    )
    st.markdown("---")
    st.markdown("**📤 Email Sending — Gmail SMTP**")
    sender_email = st.text_input(
        "📧 Your Gmail Address",
        placeholder="you@gmail.com",
        help="Must be a Gmail address. This is the From address candidates will see."
    )
    resend_api_key = st.text_input(
        "🔑 Gmail App Password",
        type="password",
        placeholder="xxxx xxxx xxxx xxxx",
        help=(
            "16-character App Password from myaccount.google.com/apppasswords. "
            "NOT your regular Gmail password. Spaces are ignored."
        )
    )
    st.markdown("---")
    st.markdown("**🗓️ Google Calendar Integration**")
    if not _CALENDAR_MODULE:
        st.warning("calendar_scheduler.py not found.", icon="⚠️")
    elif not _GOOGLE_AVAILABLE:
        st.warning(
            "Google API libraries missing. Run: pip install google-auth google-auth-oauthlib google-api-python-client",
            icon="📦",
        )
    else:
        gcal_creds_path = st.text_input(
            "📄 credentials.json path",
            value="credentials.json",
            help="Download from Google Cloud Console → APIs & Services → Credentials.",
        )
        gcal_calendar_id = st.text_input(
            "📅 Calendar ID",
            value="primary",
            help="primary for your main calendar, or paste a specific calendar ID.",
        )
        gcal_interviewer_email = st.text_input(
            "📧 Interviewer Email",
            value="",
            placeholder="interviewer@school.org",
            help="This email is added as organiser for every interview event.",
        )
    st.markdown("---")
    st.markdown("**🏛️ Premier Institution**")
    st.markdown("Detects IITs, IIMs, JNU, BITS Pilani, Oxford, Harvard, etc.")
    st.markdown("---")
    st.markdown("**📧 Pipeline Gmail Notifications**")
    notify_email = st.text_input(
        "🔔 Notify Email (Hiring Manager)",
        placeholder="manager@school.org",
        help="This email receives a notification whenever a candidate's stage changes. Uses the Gmail credentials above.",
    )
    st.markdown("---")
    st.markdown("**🚦 Pipeline Alerts**")
    stall_days = st.slider(
        "⚠️ Stall Alert (days)", 0, 14, 0,
        help="Flag candidates with no activity for this many days. Set to 0 to see all active candidates.",
    )
    cold_days = st.slider(
        "🧊 Cold Alert (days)", 0, 10, 0,
        help="Alert for active candidates not contacted in this many days. Set to 0 to monitor all.",
    )

# ─── Default gcal vars when module absent ─────────────────────────────────────
if not (_CALENDAR_MODULE and _GOOGLE_AVAILABLE):
    gcal_creds_path        = "credentials.json"
    gcal_calendar_id       = "primary"
    gcal_interviewer_email = ""

# ─── Default pipeline vars when sidebar not rendered ────────────────────────
if "notify_email" not in dir():
    notify_email = ""
if "stall_days" not in dir():
    stall_days = 0
if "cold_days" not in dir():
    cold_days = 0

# ─── Inputs ───────────────────────────────────────────────────────────────────
cl, cr = st.columns([1, 1], gap="large")

with cl:
    st.markdown(f'<p style="font-size:1.25rem;font-weight:700;color:{C["cream"]};margin:0 0 4px;">📋 Job Description</p>', unsafe_allow_html=True)
    st.markdown(f'<p style="font-size:.82rem;color:{C["silver"]};margin:0 0 10px;">Paste the full JD — the more detail, the better the semantic match.</p>', unsafe_allow_html=True)
    job_description = st.text_area(
        label="jd", label_visibility="collapsed", height=280,
        placeholder=(
            "We are hiring a Mathematics Teacher for Grades 9-12.\n\n"
            "Requirements:\n• B.Ed or equivalent teaching degree\n"
            "• CTET / teaching certification preferred\n"
            "• Minimum 2 years of classroom teaching experience\n"
            "• CBSE / ICSE curriculum design and lesson planning\n"
            "• Student assessment, grading, and feedback\n"
            "• Familiarity with Google Classroom"
        )
    )
    st.markdown(f'<p style="font-size:1.1rem;font-weight:700;color:{C["cream"]};margin:20px 0 4px;">🔒 Hard Filter Thresholds</p>', unsafe_allow_html=True)
    st.markdown(f'<p style="font-size:.82rem;color:{C["silver"]};margin:0 0 10px;">Candidates not meeting these are eliminated before ranking.</p>', unsafe_allow_html=True)
    fc1, fc2 = st.columns(2)
    with fc1:
        min_exp = st.slider("Min. Teaching Experience (yrs)", 0, 10, 1)
    with fc2:
        min_score_pct = st.slider("Min. Match Score (%)", 0, 60, 20)
    req_degree = st.checkbox("🎓 Must hold a teaching/education degree", value=True)
    req_cert   = st.checkbox("📜 Must have teaching certification / CTET", value=False)

with cr:
    st.markdown(f'<p style="font-size:1.25rem;font-weight:700;color:{C["cream"]};margin:0 0 4px;">📄 Upload Resumes</p>', unsafe_allow_html=True)
    st.markdown(f'<p style="font-size:.82rem;color:{C["silver"]};margin:0 0 10px;">PDF or TXT. Ranked by semantic match score.</p>', unsafe_allow_html=True)
    uploaded_files = st.file_uploader(
        label="resumes", label_visibility="collapsed",
        type=["pdf", "txt"], accept_multiple_files=True
    )
    if uploaded_files:
        rows_html = "".join(
            f'<div style="padding:2px 0;font-size:.84rem;">'
            f'<span style="color:{C["sky"]};font-weight:700;">#{i+1}</span>'
            f'&nbsp;<span style="color:{C["cream"]};">{f.name}</span></div>'
            for i, f in enumerate(uploaded_files)
        )
        st.markdown(
            f'<div style="background:rgba(14,165,233,.1);border:1px solid rgba(14,165,233,.3);'
            f'border-radius:12px;padding:14px 18px;margin-top:8px;">'
            f'<span style="font-weight:700;color:{C["sky"]};font-size:.95rem;">'
            f'✓ {len(uploaded_files)} file(s) ready</span>'
            f'<div style="margin-top:8px;line-height:1.9;">{rows_html}</div></div>',
            unsafe_allow_html=True
        )

# ─── Run button ───────────────────────────────────────────────────────────────
st.markdown(divider(), unsafe_allow_html=True)
_, bc, _ = st.columns([2, 1, 2])
with bc:
    run = st.button("🔍  Screen Candidates", use_container_width=True)

# ─── Screening ────────────────────────────────────────────────────────────────
if run:
    if not job_description.strip():
        st.error("⚠️  Please paste a job description.")
        st.stop()
    if not uploaded_files:
        st.error("⚠️  Upload at least one resume.")
        st.stop()
    if use_ai_just and not groq_key:
        st.info("ℹ️  No Groq key — using rule-based justification for screening. Add your free Groq key in the sidebar.")
        use_ai_just = False

    # ── TIER 1 UPGRADE 2: Run JD Bias Detection before screening ─────────
    if groq_key:
        with st.spinner("🔍 Scanning job description for bias..."):
            bias_report = detect_jd_bias(job_description, groq_key)
            st.session_state["bias_report"] = bias_report

    with st.spinner("🤖 Scoring resumes with Groq LLM + generating AI justifications..." if groq_key else "Scoring resumes locally + generating AI justifications..."):
        # TIER 1 UPGRADE 1: pass groq_api_key so screener uses LLM scoring
        screener = ResumeScreener(api_key="", groq_api_key=groq_key or "")
        config = {
            "min_experience_years":    min_exp,
            "min_similarity_score":    min_score_pct / 100.0,
            "require_teaching_degree": req_degree,
            "require_certification":   req_cert,
        }
        results = screener.screen(
            job_description, uploaded_files, config,
            use_ai_justification=use_ai_just
        )
        # ── Persist across reruns (needed for email generator & chatbot) ──
        st.session_state["screening_results"] = results
        st.session_state["screening_jd"]      = job_description
        # Build RAG index over all candidate profiles for semantic search
        if _RAG_AVAILABLE:
            try:
                st.session_state["rag_index"] = build_index(results)
            except Exception:
                st.session_state["rag_index"] = None
        else:
            st.session_state["rag_index"] = None

        # ── Initialise pipeline tracker ──────────────────────────────────
        init_pipeline(results, st.session_state)

    # ── Stats ─────────────────────────────────────────────────────────────
    st.markdown(divider(), unsafe_allow_html=True)
    st.markdown(f'<p style="font-size:1.3rem;font-weight:700;color:{C["cream"]};margin:0 0 16px;">📊 Screening Summary</p>', unsafe_allow_html=True)

    qualified = [r for r in results if not r["filtered"]]
    filtered  = [r for r in results if r["filtered"]]
    top_score = max((r["score"] for r in qualified), default=0)
    avg_score = (sum(r["score"] for r in qualified) / len(qualified)) if qualified else 0

    s1, s2, s3, s4, s5 = st.columns(5)
    for col, val, lbl, fg, bg_col in [
        (s1, len(results),         "Total Screened", C["cream"],  "rgba(14,165,233,.12)"),
        (s2, len(qualified),       "Qualified",       C["sage"],   "rgba(16,185,129,.12)"),
        (s3, len(filtered),        "Filtered Out",    C["rose"],   "rgba(244,63,94,.12)"),
        (s4, f"{top_score:.0f}%",  "Top Score",       C["gold"],   "rgba(245,158,11,.12)"),
        (s5, f"{avg_score:.0f}%",  "Avg Score",       C["silver"], "rgba(148,163,184,.12)"),
    ]:
        col.markdown(
            f'<div style="background:{bg_col};border:1px solid {fg}33;'
            f'border-radius:14px;padding:20px 18px;text-align:center;">'
            f'<div style="font-size:2.6rem;font-weight:900;line-height:1;color:{fg};">{val}</div>'
            f'<div style="font-size:.7rem;color:{C["silver"]};text-transform:uppercase;'
            f'letter-spacing:1px;margin-top:4px;">{lbl}</div></div>',
            unsafe_allow_html=True
        )

    st.markdown("<br>", unsafe_allow_html=True)

    methods = set(r.get("score_method", "") for r in results)
    if "groq_llm" in methods:
        st.success("🤖 **LLM scoring active** — scores computed by Groq / Llama 3.3 70B with reasoning, strengths & red flags.")
    elif "sentence_transformer" in methods:
        st.success("✅ **Semantic scoring active** — scores computed locally via `sentence-transformers/all-MiniLM-L6-v2`.")
    elif "tfidf" in methods:
        st.warning("⚠️ **TF-IDF scoring** — `sentence-transformers` not installed. Run `pip install sentence-transformers` for semantic scores.")
    else:
        st.error("❌ **Scoring failed** — check that `sentence-transformers` or `scikit-learn` is installed.")

    just_source = "🤖 AI (Groq/Llama)" if use_ai_just and groq_key else "📝 Rule-based"
    st.info(f"**Justification source:** {just_source}")

    # ── TIER 1 UPGRADE 2: JD Bias Detection Report ────────────────────────
    bias_report = st.session_state.get("bias_report")
    if bias_report and bias_report.get("checked"):
        risk       = bias_report.get("overall_risk", "Unknown")
        issues     = bias_report.get("issues", [])
        summary    = bias_report.get("summary", "")
        risk_color = {"Low": "#10B981", "Medium": "#F59E0B", "High": "#EF4444"}.get(risk, "#94A3B8")
        risk_emoji = {"Low": "✅", "Medium": "⚠️", "High": "🚨"}.get(risk, "❓")

        st.markdown(divider(), unsafe_allow_html=True)
        st.markdown(
            f'<p style="font-size:1.2rem;font-weight:700;color:{C["cream"]};margin:0 0 10px;">'
            f'🏳️ JD Bias Detection Report</p>',
            unsafe_allow_html=True
        )
        st.markdown(
            f'<div style="background:rgba(0,0,0,.25);border:1.5px solid {risk_color}55;'
            f'border-radius:14px;padding:18px 22px;margin-bottom:12px;">'
            f'<span style="font-size:1.1rem;font-weight:800;color:{risk_color};">'
            f'{risk_emoji} Bias Risk: {risk}</span>'
            f'<p style="color:#D1DCF0;margin:8px 0 0;font-size:.9rem;">{summary}</p>'
            f'</div>',
            unsafe_allow_html=True
        )
        if issues:
            for issue in issues:
                itype      = issue.get("type", "Issue")
                example    = issue.get("example", "")
                suggestion = issue.get("suggestion", "")
                st.markdown(
                    f'<div style="background:rgba(245,158,11,.08);border-left:3px solid #F59E0B;'
                    f'border-radius:0 10px 10px 0;padding:10px 16px;margin-bottom:8px;">'
                    f'<span style="color:#F59E0B;font-weight:700;font-size:.85rem;">{itype}</span><br>'
                    f'<span style="color:#E8EFF8;font-size:.85rem;">Found: <em>"{example}"</em></span><br>'
                    f'<span style="color:#10B981;font-size:.85rem;">Suggestion: {suggestion}</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )
        else:
            st.markdown(
                f'<p style="color:#10B981;font-size:.88rem;">No bias issues detected in this job description.</p>',
                unsafe_allow_html=True
            )
    elif groq_key:
        st.info("ℹ️ Bias detection ran but could not produce a report. Check your Groq key.")

    # ══════════════════════════════════════════════════════════════════════
    # VISUAL ANALYTICS DASHBOARD
    # ══════════════════════════════════════════════════════════════════════

    st.markdown(divider(), unsafe_allow_html=True)
    st.markdown(
        f'<p style="font-size:1.4rem;font-weight:800;color:{C["cream"]};margin:0 0 6px;'
        f'font-family:\'DM Serif Display\',Georgia,serif;">📊 Visual Analytics Dashboard</p>'
        f'<p style="font-size:.83rem;color:{C["silver"]};margin:0 0 24px;">'
        f'Interactive charts to help you make faster, smarter hiring decisions.</p>',
        unsafe_allow_html=True
    )

    tab_overview, tab_comparison, tab_gaps, tab_breakdown = st.tabs([
        "📈 Overview", "🕸 Candidate Comparison", "🔍 JD Gap Analysis", "📋 Breakdown",
    ])

    with tab_overview:
        col_a, col_b = st.columns(2, gap="medium")
        with col_a:
            st.plotly_chart(chart_score_bar(results), use_container_width=True)
        with col_b:
            st.plotly_chart(chart_score_distribution(results), use_container_width=True)
        col_c, col_d = st.columns(2, gap="medium")
        with col_c:
            st.plotly_chart(chart_score_vs_experience(results), use_container_width=True)
        with col_d:
            st.plotly_chart(chart_experience_distribution(results), use_container_width=True)

    with tab_comparison:
        st.markdown(
            f'<p style="font-size:.82rem;color:{C["silver"]};margin:0 0 16px;">'
            f'Radar chart compares top 3 qualified candidates across 16 dimensions. '
            f'Heatmap gives a full-pool overview at a glance.</p>',
            unsafe_allow_html=True
        )
        st.plotly_chart(chart_radar_top3(results), use_container_width=True)
        st.plotly_chart(chart_candidate_heatmap(results), use_container_width=True)

    with tab_gaps:
        st.markdown(
            f'<p style="font-size:.82rem;color:{C["silver"]};margin:0 0 16px;">'
            f'See which JD keywords your talent pool covers — and where gaps exist.</p>',
            unsafe_allow_html=True
        )
        st.plotly_chart(chart_keyword_gap(results), use_container_width=True)
        st.plotly_chart(chart_skills_frequency(results), use_container_width=True)

    with tab_breakdown:
        st.plotly_chart(chart_credentials_donut(results), use_container_width=True)

    # ══════════════════════════════════════════════════════════════════════
    # QUALIFIED CANDIDATES
    # ══════════════════════════════════════════════════════════════════════
    if qualified:
        st.markdown(
            f'<p style="font-size:1.3rem;font-weight:700;color:{C["sage"]};margin:28px 0 4px;">'
            f'✅ Qualified Candidates — Ranked by Match Score</p>'
            f'<p style="font-size:.82rem;color:{C["silver"]};margin:0 0 16px;">'
            f'Rank 1 = highest semantic similarity. Scored locally via sentence-transformers.</p>',
            unsafe_allow_html=True
        )

        for res in qualified:
            rank  = res["rank"]
            score = res["score"]

            if rank == 1:
                card_bg, border_col = "rgba(245,158,11,.1)", C["gold"]
                rank_icon = "🥇"; pill_bg, pill_fg = "rgba(245,158,11,.2)", C["gold"]
                bar_color = "linear-gradient(90deg," + C["sage"] + "," + C["gold"] + ")"
            elif rank == 2:
                card_bg, border_col = "rgba(148,163,184,.08)", C["silver"]
                rank_icon = "🥈"; pill_bg, pill_fg = "rgba(16,185,129,.2)", C["sage"]
                bar_color = "linear-gradient(90deg," + C["sage"] + "," + C["sky"] + ")"
            elif rank == 3:
                card_bg, border_col = "rgba(217,119,6,.08)", C["bronze"]
                rank_icon = "🥉"; pill_bg, pill_fg = "rgba(16,185,129,.2)", C["sage"]
                bar_color = "linear-gradient(90deg," + C["sage"] + "," + C["sky"] + ")"
            else:
                card_bg, border_col = "rgba(14,165,233,.06)", C["sky"]
                rank_icon = f"#{rank}"
                if score >= 50:   pill_bg, pill_fg, bar_color = "rgba(16,185,129,.2)", C["sage"], "linear-gradient(90deg," + C["sage"] + "," + C["sky"] + ")"
                elif score >= 25: pill_bg, pill_fg, bar_color = "rgba(245,158,11,.2)", C["amber"], "linear-gradient(90deg," + C["amber"] + ",#FCD34D)"
                else:             pill_bg, pill_fg, bar_color = "rgba(244,63,94,.2)", C["rose"], "linear-gradient(90deg," + C["rose"] + ",#F87171)"

            t_html  = green_tags(res.get("tags", []))
            s_pill  = score_pill(score, pill_bg, pill_fg)
            s_bar   = score_bar(score, bar_color)
            m_badge = method_badge(res.get("score_method", ""))
            p_badge = premier_badge(res.get("is_premier", False))
            exp_str = str(res.get("exp", 0)) + " yr exp"
            matched = kw_tags(res.get("matched_kw", [])[:6], matched=True)
            missing = kw_tags(res.get("missing_kw", [])[:4], matched=False)
            jst       = _clean_text(res.get("justification", ""))
            cand_name = _clean_text(res.get("name", "Unknown"))

            # Pre-compute all conditional HTML to avoid nested quote conflicts in f-string
            deg_html = '<span style="font-size:.75rem;color:#6EE7B7;">&#10003; Degree</span>' if res.get("has_deg") else ""
            cert_html_q = '<span style="font-size:.75rem;color:#6EE7B7;">&#10003; Certified</span>' if res.get("has_cert") else ""
            matched_none = f'<span style="font-size:.72rem;color:{C["silver"]};">None detected</span>'
            matched_display = matched if matched else matched_none
            gaps_html = (
                '<div style="margin-top:6px;display:flex;flex-wrap:wrap;gap:6px;align-items:center;">'
                f'<span style="font-size:.75rem;font-weight:600;color:{C["rose"]};">JD Gaps:</span>'
                + missing +
                '</div>'
            ) if missing else ""

            html_content = "".join([
                f'<div style="background:{card_bg};border-left:5px solid {border_col};',
                f'border-radius:0 16px 16px 0;padding:22px 28px;margin-bottom:20px;',
                f'box-shadow:0 4px 24px rgba(0,0,0,.25);backdrop-filter:blur(8px);">',
                f'<div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:12px;">',
                f'<div style="display:flex;align-items:center;gap:18px;">',
                f'<div style="font-size:2.2rem;font-weight:900;line-height:1;min-width:44px;">{rank_icon}</div>',
                f'<div>',
                f'<p style="font-size:1.18rem;font-weight:700;color:#E8EFF8;margin:0 0 5px;',
                f'font-family:\'DM Serif Display\',Georgia,serif;">{cand_name}</p>',
                f'<div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:8px;">',
                f'{m_badge} {p_badge} <span style="font-size:.75rem;color:{C["silver"]};">{exp_str}</span> ',
                f'{deg_html} {cert_html_q}',
                f'</div>',
                f'<div style="display:flex;flex-wrap:wrap;gap:4px;">{t_html}</div>',
                f'</div>',
                f'</div>',
                f'<div style="text-align:right;min-width:155px;">',
                f'{s_pill} {s_bar}',
                f'</div>',
                f'</div>',
                f'<div style="margin-top:10px;display:flex;flex-wrap:wrap;gap:6px;align-items:center;">',
                f'<span style="font-size:.75rem;font-weight:600;color:{C["sage"]};">JD Matched:</span> ',
                f'{matched_display}',
                f'</div>',
                f'{gaps_html}',
                f'<div style="margin-top:14px;padding:14px 18px;background:rgba(14,165,233,.08);',
                f'border-left:4px solid {C["sky"]};border-radius:10px;">',
                f'<span style="font-size:.88rem;color:{C["cream"]};line-height:1.75;">&#128172; {jst}</span>',
                f'</div>',
                f'</div>'
            ])
            st.markdown(html_content, unsafe_allow_html=True)

            # ── ATS Score Breakdown ───────────────────────────────────────────
            ats = res.get("ats_breakdown", {})
            if ats:
                with st.expander(f"📊 ATS Score Breakdown for {res.get('name','?')}", expanded=False):
                    st.markdown(render_ats_breakdown(ats), unsafe_allow_html=True)
                    kw_data = ats.get("keywords", {}).get("detail", {})
                    if isinstance(kw_data, dict):
                        matched_kw_ats = kw_data.get("matched", [])
                        missing_kw_ats = kw_data.get("missing", [])
                        if matched_kw_ats:
                            matched_str = ', '.join(matched_kw_ats)
                            st.markdown(
                                f'<p style="font-size:.75rem;color:#10B981;margin:8px 0 2px;font-weight:600;">✓ JD Keywords Matched: {matched_str}</p>',
                                unsafe_allow_html=True
                            )
                        if missing_kw_ats:
                            missing_str = ', '.join(missing_kw_ats)
                            st.markdown(
                                f'<p style="font-size:.75rem;color:#F43F5E;margin:2px 0 4px;font-weight:600;">✗ Keywords Missing: {missing_str}</p>',
                                unsafe_allow_html=True
                            )
                    weighted = ats.get("weighted_ats", 0)
                    overall  = ats.get("overall_score", 0)
                    st.markdown(
                        f'<p style="font-size:.72rem;color:#64748B;margin-top:8px;">' +
                        f'ATS Weighted Score: <strong style="color:#10B981;">{weighted:.0f}%</strong> &nbsp;·&nbsp; ' +
                        f'Overall Match Score: <strong style="color:#0EA5E9;">{overall:.0f}%</strong></p>',
                        unsafe_allow_html=True
                    )

    else:
        st.warning("No candidates passed the hard filters. Try lowering thresholds.")

    # ══════════════════════════════════════════════════════════════════════
    # FILTERED CANDIDATES
    # ══════════════════════════════════════════════════════════════════════
    if filtered:
        with st.expander(f"❌  {len(filtered)} candidate(s) filtered out — click to expand"):
            st.markdown(
                f'<p style="font-size:.82rem;color:{C["silver"]};margin:0 0 12px;">'
                f'Sorted by: Premier Institution first, then by match score (highest to lowest).</p>',
                unsafe_allow_html=True
            )
            for res in filtered:
                t_html  = rose_tags(res.get("tags", []))
                matched = kw_tags(res.get("matched_kw", [])[:4], matched=True)
                s_pill  = score_pill(res["score"], "rgba(244,63,94,.15)", C["rose"])
                s_bar   = score_bar(res["score"], "linear-gradient(90deg," + C["rose"] + ",#F87171)")
                p_badge = premier_badge(res.get("is_premier", False))
                jst     = _clean_text(res.get("justification", ""))
                cand_name = _clean_text(res.get("name", "Unknown"))
                exp_str = str(res.get("exp", 0)) + " yr exp"

                # Pre-compute all conditional HTML snippets BEFORE the f-string
                # to avoid nested quote conflicts that cause raw HTML to appear in UI
                degree_html = (
                    '<span style="font-size:.72rem;color:#6EE7B7;">&#10003; Degree</span>'
                    if res.get("has_deg") else
                    '<span style="font-size:.72rem;color:#FDA4AF;">&#10007; No Degree</span>'
                )
                cert_html = (
                    '<span style="font-size:.72rem;color:#6EE7B7;">&#10003; Certified</span>'
                    if res.get("has_cert") else ""
                )
                matched_row_html = (
                    '<div style="margin-top:10px;display:flex;flex-wrap:wrap;gap:4px;align-items:center;">'
                    '<span style="font-size:.72rem;font-weight:600;color:#6EE7B7;">Partial matches:</span>'
                    + matched +
                    '</div>'
                ) if matched else ""

                html_content = "".join([
                    f'<div style="background:rgba(244,63,94,.06);border-left:5px solid {C["rose"]};',
                    f'border-radius:0 16px 16px 0;padding:22px 28px;margin-bottom:14px;',
                    f'box-shadow:0 4px 16px rgba(0,0,0,.2);">',
                    f'<div style="display:flex;align-items:flex-start;justify-content:space-between;flex-wrap:wrap;gap:12px;">',
                    f'<div style="display:flex;align-items:center;gap:16px;">',
                    f'<div style="font-size:1.3rem;font-weight:900;color:{C["rose"]};',
                    f'background:rgba(244,63,94,.15);border-radius:50%;width:36px;height:36px;',
                    f'display:flex;align-items:center;justify-content:center;flex-shrink:0;">&#10005;</div>',
                    f'<div>',
                    f'<p style="font-size:1.1rem;font-weight:700;color:{C["cream"]};margin:0 0 5px;',
                    f'font-family:\'DM Serif Display\',Georgia,serif;">{cand_name}</p>',
                    f'<div style="display:flex;align-items:center;gap:6px;flex-wrap:wrap;margin-bottom:7px;">',
                    f'{p_badge} <span style="font-size:.75rem;color:{C["silver"]};">File #{res["upload_pos"]}</span> ',
                    f'<span style="font-size:.75rem;color:{C["silver"]};">{exp_str}</span> ',
                    f'{degree_html} {cert_html}',
                    f'</div>',
                    f'<div style="display:flex;flex-wrap:wrap;gap:4px;">{t_html}</div>',
                    f'</div>',
                    f'</div>',
                    f'<div style="text-align:right;min-width:155px;">',
                    f'{s_pill} {s_bar}',
                    f'</div>',
                    f'</div>',
                    f'{matched_row_html}',
                    f'<div style="margin-top:14px;padding:14px 18px;background:rgba(244,63,94,.08);',
                    f'border-left:4px solid {C["rose"]};border-radius:10px;">',
                    f'<span style="font-size:.87rem;color:{C["cream"]};line-height:1.75;">&#128683; {jst}</span>',
                    f'</div>',
                    f'</div>'
                ])
                st.markdown(html_content, unsafe_allow_html=True)

                # ── ATS Score Breakdown for filtered candidates ───────────────
                ats_f = res.get("ats_breakdown", {})
                if ats_f:
                    with st.expander(f"📊 ATS Score Breakdown for {res.get('name','?')}", expanded=False):
                        st.markdown(render_ats_breakdown(ats_f), unsafe_allow_html=True)
                        kw_data_f = ats_f.get("keywords", {}).get("detail", {})
                        if isinstance(kw_data_f, dict):
                            matched_kw_f = kw_data_f.get("matched", [])
                            missing_kw_f = kw_data_f.get("missing", [])
                            if matched_kw_f:
                                st.markdown(
                                    f'<p style="font-size:.75rem;color:#10B981;margin:8px 0 2px;font-weight:600;">✓ JD Keywords Matched: {", ".join(matched_kw_f)}</p>',
                                    unsafe_allow_html=True
                                )
                            if missing_kw_f:
                                st.markdown(
                                    f'<p style="font-size:.75rem;color:#F43F5E;margin:2px 0 4px;font-weight:600;">✗ Keywords Missing: {", ".join(missing_kw_f)}</p>',
                                    unsafe_allow_html=True
                                )
    # ══════════════════════════════════════════════════════════════════════
    st.markdown("<br>", unsafe_allow_html=True)

    def clean_cell(text):
        if not isinstance(text, str): text = str(text)
        text = re.sub(r'<[^>]+>', '', text)  # strip HTML
        return (text.replace("\u2014","-").replace("\u2013","-")
                    .replace("\u2018","'").replace("\u2019","'")
                    .replace("\u201C",'"').replace("\u201D",'"')
                    .replace("\u2026","...").replace("\u00A0"," "))

    def hdr_style(ws, hex_col, cols):
        fill   = PatternFill("solid", fgColor=hex_col)
        font   = Font(bold=True, color="FFFFFF", size=11)
        thin   = Side(style="thin", color="D0D0D0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for ci, _ in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci)
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
        ws.row_dimensions[1].height = 28

    def row_style(ws, ri, nc, alt=False):
        thin   = Side(style="thin", color="E8E8E8")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        fill   = PatternFill("solid", fgColor="F7F7F7") if alt else PatternFill("solid", fgColor="FFFFFF")
        for ci in range(1, nc+1):
            c = ws.cell(row=ri, column=ci)
            c.fill = fill; c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[ri].height = 22

    def col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def freeze_filter(ws, n):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(n)}1"

    wb = Workbook()

    # ── Sheet 1: Qualified ─────────────────────────────────────────────────
    ws_q = wb.active
    ws_q.title = "Qualified Candidates"
    q_hdrs = ["Rank","Name","Match Score (%)","ATS Score (%)","Education ATS","Experience ATS","Skills ATS","Keywords ATS","Certification ATS","Experience (yrs)",
              "Has Degree","Has Certification","Premier Institution",
              "Skills","JD Keywords Matched","JD Keywords Missing","Justification"]
    ws_q.append(q_hdrs)
    hdr_style(ws_q, "0F766E", q_hdrs)

    for i, r in enumerate(qualified):
        row_style(ws_q, i+2, len(q_hdrs), alt=(i%2==1))
        ats = r.get("ats_breakdown", {})
        ws_q.append([
            r.get("rank",""),
            clean_cell(r.get("name","")),
            round(r.get("score",0), 1),
            round(ats.get("weighted_ats", 0), 1),
            round(ats.get("education", {}).get("score", 0), 1),
            round(ats.get("experience", {}).get("score", 0), 1),
            round(ats.get("skills", {}).get("score", 0), 1),
            round(ats.get("keywords", {}).get("score", 0), 1),
            round(ats.get("certification", {}).get("score", 0), 1),
            r.get("exp",0),
            "Yes" if r.get("has_deg") else "No",
            "Yes" if r.get("has_cert") else "No",
            "Yes" if r.get("is_premier") else "No",
            clean_cell(", ".join(r.get("tags",[]))),
            clean_cell(", ".join(r.get("matched_kw",[]))),
            clean_cell(", ".join(r.get("missing_kw",[])))[:120],
            clean_cell(r.get("justification","")),
        ])
        sc = ws_q.cell(row=i+2, column=3)
        s  = r.get("score",0)
        if s>=50: sc.fill=PatternFill("solid",fgColor="D4EDE8"); sc.font=Font(bold=True,color="1F5C50")
        elif s>=25: sc.fill=PatternFill("solid",fgColor="FEF3C7"); sc.font=Font(bold=True,color="92400E")
        else: sc.fill=PatternFill("solid",fgColor="FAD9D9"); sc.font=Font(bold=True,color="8B1A1A")
        # ATS score cell (column 4)
        ats_sc = ws_q.cell(row=i+2, column=4)
        ats_val = ats.get("weighted_ats", 0)
        if ats_val>=60: ats_sc.fill=PatternFill("solid",fgColor="D4EDE8"); ats_sc.font=Font(bold=True,color="1F5C50")
        elif ats_val>=35: ats_sc.fill=PatternFill("solid",fgColor="FEF3C7"); ats_sc.font=Font(bold=True,color="92400E")
        else: ats_sc.fill=PatternFill("solid",fgColor="FAD9D9"); ats_sc.font=Font(bold=True,color="8B1A1A")
        pc = ws_q.cell(row=i+2, column=13)  # Premier Institution now at col 13
        if r.get("is_premier"):
            pc.fill = PatternFill("solid", fgColor="EDE9FE")
            pc.font = Font(bold=True, color="5B21B6")

    col_widths(ws_q, [7,22,15,13,13,13,12,13,16,13,12,17,18,30,30,30,70])
    freeze_filter(ws_q, len(q_hdrs))
    if qualified:
        avg = sum(r["score"] for r in qualified)/len(qualified)
        ws_q.append([])
        si = len(qualified)+3
        ws_q.cell(si,1).value="SUMMARY"; ws_q.cell(si,1).font=Font(bold=True,color="0F766E")
        ws_q.cell(si,2).value=f"{len(qualified)} qualified"
        ws_q.cell(si,3).value=f"Avg: {avg:.1f}%"; ws_q.cell(si,3).font=Font(bold=True)

    # ── Sheet 2: Filtered Out — ranked: Premier Institution first, then score ──
    ws_f = wb.create_sheet("Filtered Out")
    f_hdrs = ["Rank (Premier→Score)","Name","Match Score (%)","ATS Score (%)","Education ATS","Experience ATS","Skills ATS","Keywords ATS","Certification ATS","Experience (yrs)",
              "Has Degree","Has Certification","Premier Institution",
              "Skills","JD Partial Matches","Filter Reason","Justification"]
    ws_f.append(f_hdrs)
    hdr_style(ws_f, "B83232", f_hdrs)

    # filtered is already sorted: premier first, then by score desc (from backend)
    for i, r in enumerate(filtered):
        row_style(ws_f, i+2, len(f_hdrs), alt=(i%2==1))
        reason = clean_cell("; ".join(r.get("fail_reasons",[])))
        ats_f = r.get("ats_breakdown", {})
        ws_f.append([
            i+1,  # rank within filtered sheet
            clean_cell(r.get("name","")),
            round(r.get("score",0), 1),
            round(ats_f.get("weighted_ats", 0), 1),
            round(ats_f.get("education", {}).get("score", 0), 1),
            round(ats_f.get("experience", {}).get("score", 0), 1),
            round(ats_f.get("skills", {}).get("score", 0), 1),
            round(ats_f.get("keywords", {}).get("score", 0), 1),
            round(ats_f.get("certification", {}).get("score", 0), 1),
            r.get("exp",0),
            "Yes" if r.get("has_deg") else "No",
            "Yes" if r.get("has_cert") else "No",
            "Yes" if r.get("is_premier") else "No",
            clean_cell(", ".join(r.get("tags",[]))),
            clean_cell(", ".join(r.get("matched_kw",[]))),
            reason,
            clean_cell(r.get("justification","")),
        ])
        ws_f.cell(i+2,16).fill=PatternFill("solid",fgColor="FFF0F0")
        ws_f.cell(i+2,16).font=Font(color="8B1A1A")
        pc = ws_f.cell(row=i+2, column=13)
        if r.get("is_premier"):
            pc.fill = PatternFill("solid", fgColor="EDE9FE")
            pc.font = Font(bold=True, color="5B21B6")
        # Gold row highlight for premier rows
        if r.get("is_premier"):
            rank_cell = ws_f.cell(row=i+2, column=1)
            rank_cell.fill = PatternFill("solid", fgColor="F5F0FF")
            rank_cell.font = Font(bold=True, color="5B21B6")

    col_widths(ws_f, [18,22,15,13,13,13,12,13,16,13,12,17,18,28,28,40,70])
    freeze_filter(ws_f, len(f_hdrs))
    if filtered:
        ws_f.append([])
        si = len(filtered)+3
        ws_f.cell(si,1).value="SUMMARY"; ws_f.cell(si,1).font=Font(bold=True,color="B83232")
        ws_f.cell(si,2).value=f"{len(filtered)} filtered out"
        premier_filtered = sum(1 for r in filtered if r.get("is_premier"))
        ws_f.cell(si,3).value=f"{premier_filtered} from premier institutions"
        ws_f.cell(si,3).font=Font(bold=True,color="5B21B6")

    # ── Sheet 3: Overview ──────────────────────────────────────────────────
    ws_o = wb.create_sheet("Overview")
    ws_o.column_dimensions["A"].width = 32
    ws_o.column_dimensions["B"].width = 22
    ws_o["A1"]="EduHire Screening Report"; ws_o["A1"].font=Font(bold=True,size=16,color="0D1B2A")
    ws_o["A2"]="Local sentence-transformers scoring + Groq AI chatbot (Llama/Mixtral)"
    ws_o["A2"].font=Font(size=10,color="888888")
    ws_o.row_dimensions[1].height=30
    premier_qual  = sum(1 for r in qualified if r.get("is_premier"))
    premier_filt  = sum(1 for r in filtered  if r.get("is_premier"))
    scoring_m = next((r.get("score_method","") for r in results), "")
    scoring_label = "sentence-transformers (local)" if scoring_m=="sentence_transformer" else "TF-IDF (local)" if scoring_m=="tfidf" else "unknown"
    metrics = [
        ("",""),("SCREENING SUMMARY",""),
        ("Total Resumes Screened",len(results)),
        ("Qualified Candidates",len(qualified)),
        ("Filtered Out",len(filtered)),
        ("Filter Rate",f"{len(filtered)/len(results)*100:.1f}%" if results else "0%"),
        ("",""),("SCORE ANALYTICS",""),
        ("Scoring Method", scoring_label),
        ("Top Match Score",f"{max((r['score'] for r in qualified),default=0):.1f}%"),
        ("Avg Match Score (Qualified)",f"{sum(r['score'] for r in qualified)/len(qualified):.1f}%" if qualified else "N/A"),
        ("Lowest Passing Score",f"{min((r['score'] for r in qualified),default=0):.1f}%"),
        ("",""),("CANDIDATE QUALITY",""),
        ("With Education Degree",sum(1 for r in qualified if r.get("has_deg"))),
        ("With Certification",sum(1 for r in qualified if r.get("has_cert"))),
        ("From Premier Institution (Qualified)", premier_qual),
        ("From Premier Institution (Filtered)", premier_filt),
        ("Avg Experience (Qualified, yrs)",f"{sum(r.get('exp',0) for r in qualified)/len(qualified):.1f}" if qualified else "N/A"),
        ("",""),("FILTERING NOTES",""),
        ("Filtered Out — Premier First","Yes — Filtered Out sheet ranks premier institution candidates first, then by score"),
        ("Premier Institution Detection","Strict: IITs, IIMs, JNU, BITS Pilani, Oxford, Harvard, etc. only"),
    ]
    for ri,(label,value) in enumerate(metrics,start=4):
        ws_o.cell(ri,1).value=label; ws_o.cell(ri,2).value=value
        if label in ("SCREENING SUMMARY","SCORE ANALYTICS","CANDIDATE QUALITY","FILTERING NOTES"):
            ws_o.cell(ri,1).font=Font(bold=True,size=11,color="0F766E")
            ws_o.cell(ri,1).fill=PatternFill("solid",fgColor="F0FDF9")
        elif label:
            ws_o.cell(ri,1).font=Font(size=11,color="555555")
            ws_o.cell(ri,2).font=Font(bold=True,size=12)
    wb.move_sheet("Overview", offset=-(len(wb.sheetnames)-1))

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_buf.seek(0)

    st.download_button(
        label="⬇️  Download Excel Report (3 Sheets: Overview + Qualified + Filtered)",
        data=excel_buf.getvalue(),
        file_name="eduhire_screening_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ══════════════════════════════════════════════════════════════════════════════
# EMAIL DRAFT GENERATOR
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(divider(), unsafe_allow_html=True)

st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(16,185,129,.12),rgba(14,165,233,.10));
     border:1px solid rgba(16,185,129,.3);border-radius:16px;
     padding:28px 36px;margin-bottom:24px;">
  <div style="display:flex;align-items:center;gap:16px;margin-bottom:10px;">
    <div style="font-size:2rem;">📧</div>
    <div>
      <h2 style="font-size:1.5rem;font-weight:800;color:#E8EFF8;margin:0;
        font-family:'DM Serif Display',Georgia,serif;">Auto Email Draft Generator</h2>
      <p style="font-size:.83rem;color:{C['silver']};margin:4px 0 0;">
        Instantly generate personalised shortlist &amp; rejection emails for every candidate.
        Edit before copying — powered by Groq AI or smart templates.
      </p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if "screening_results" in st.session_state and st.session_state.screening_results:
    _email_results = st.session_state.screening_results
    _email_jd      = st.session_state.get("screening_jd", "")

    # ── Session state for generated emails ────────────────────────────────
    if "email_drafts" not in st.session_state:
        st.session_state.email_drafts = {}

    qualified_for_email = [r for r in _email_results if not r["filtered"]]
    filtered_for_email  = [r for r in _email_results if r["filtered"]]

    email_tab_shortlist, email_tab_rejection, email_tab_bulk = st.tabs([
        f"✅ Shortlist Emails ({len(qualified_for_email)})",
        f"❌ Rejection Emails ({len(filtered_for_email)})",
        "📦 Bulk Download All",
    ])

    def _render_email_card(res, tab_key_prefix):
        cand_key    = f"{tab_key_prefix}_{res.get('name','?')}_{res.get('upload_pos',0)}"
        name        = res.get("name", "Candidate")
        is_filtered = res.get("filtered", False)

        btn_label = f"{'📧' if not is_filtered else '📩'} Generate Email for {name}"
        col_btn, col_status = st.columns([3, 2])
        with col_btn:
            if st.button(btn_label, key=f"gen_{cand_key}", use_container_width=True):
                with st.spinner(f"Drafting email for {name}..."):
                    draft = generate_email_draft(
                        candidate=res,
                        job_description=_email_jd,
                        school_name=school_name,
                        hr_name=hr_name,
                        groq_api_key=groq_key,
                        groq_model=groq_model,
                    )
                st.session_state.email_drafts[cand_key] = draft
                st.rerun()

        with col_status:
            if cand_key in st.session_state.email_drafts:
                src = "🤖 AI-generated" if groq_key else "📝 Template"
                st.markdown(
                    f'<div style="padding:8px 12px;background:rgba(16,185,129,.12);'
                    f'border:1px solid rgba(16,185,129,.3);border-radius:10px;font-size:.78rem;'
                    f'color:#6EE7B7;text-align:center;">{src}</div>',
                    unsafe_allow_html=True
                )

        # Show editable draft if generated
        if cand_key in st.session_state.email_drafts:
            draft = st.session_state.email_drafts[cand_key]

            st.markdown(
                f'<p style="font-size:.8rem;font-weight:600;color:{C["sky"]};margin:10px 0 2px;">Subject Line</p>',
                unsafe_allow_html=True
            )
            edited_subject = st.text_input(
                label="subject",
                label_visibility="collapsed",
                value=draft["subject"],
                key=f"subj_{cand_key}",
            )

            st.markdown(
                f'<p style="font-size:.8rem;font-weight:600;color:{C["sky"]};margin:10px 0 2px;">Email Body</p>',
                unsafe_allow_html=True
            )
            edited_body = st.text_area(
                label="body",
                label_visibility="collapsed",
                value=draft["body"],
                height=280,
                key=f"body_{cand_key}",
            )

            # Save edits back to session state
            st.session_state.email_drafts[cand_key]["subject"] = edited_subject
            st.session_state.email_drafts[cand_key]["body"]    = edited_body

            full_email_text = f"Subject: {edited_subject}\n\n{edited_body}"

            # ── Send Email Section ────────────────────────────────────────────
            st.markdown(
                f'<p style="font-size:.8rem;font-weight:600;color:{C["sky"]};margin:12px 0 2px;">📨 Send Email via Gmail</p>',
                unsafe_allow_html=True
            )
            recipient_email_input = st.text_input(
                label="recipient_email",
                label_visibility="collapsed",
                placeholder=f"Enter {name}'s email address to send directly",
                key=f"recip_{cand_key}",
            )

            send_col, dl_col, status_col = st.columns([2, 1, 3])
            with send_col:
                can_send = bool(resend_api_key and sender_email and recipient_email_input)
                send_clicked = st.button(
                    "📤 Send Email",
                    key=f"send_{cand_key}",
                    use_container_width=True,
                    disabled=not can_send,
                    help="Add Gmail address + App Password in sidebar, then enter recipient email above." if not can_send else "Send now"
                )
                if send_clicked and can_send:
                    with st.spinner(f"Sending to {recipient_email_input}..."):
                        result = send_email(
                            gmail_app_password=resend_api_key,
                            sender_email=sender_email,
                            recipient_email=recipient_email_input,
                            subject=edited_subject,
                            body=edited_body,
                            sender_name=hr_name or "EduHire Screener",
                        )
                    st.session_state[f"send_status_{cand_key}"] = result

            with dl_col:
                st.download_button(
                    label="⬇️ .txt",
                    data=full_email_text,
                    file_name=f"email_{name.replace(' ', '_')}.txt",
                    mime="text/plain",
                    key=f"dl_{cand_key}",
                )

            with status_col:
                send_status = st.session_state.get(f"send_status_{cand_key}")
                if send_status:
                    color = "#10B981" if send_status["success"] else "#F43F5E"
                    st.markdown(
                        f'<div style="padding:8px 12px;background:rgba(0,0,0,.2);' +
                        f'border:1px solid {color}44;border-radius:10px;font-size:.78rem;' +
                        f'color:{color};">{send_status["message"]}</div>',
                        unsafe_allow_html=True
                    )
                elif not resend_api_key or not sender_email:
                    st.markdown(
                        f'<div style="padding:8px 12px;background:rgba(245,158,11,.08);' +
                        f'border:1px solid rgba(245,158,11,.3);border-radius:10px;font-size:.75rem;' +
                        f'color:#F59E0B;">Add Gmail address + App Password in sidebar to send directly.</div>',
                        unsafe_allow_html=True
                    )

            st.markdown(
                "<hr style='border:none;height:1px;background:rgba(30,58,95,.5);margin:18px 0;'>",
                unsafe_allow_html=True
            )

    # ── Shortlist emails tab ──────────────────────────────────────────────
    with email_tab_shortlist:
        if not qualified_for_email:
            st.info("No qualified candidates to generate shortlist emails for.")
        else:
            st.markdown(
                f'<p style="font-size:.83rem;color:{C["silver"]};margin:0 0 16px;">'
                f'Generate personalised shortlist / invitation emails for each qualified candidate.</p>',
                unsafe_allow_html=True
            )
            # Generate all shortlist emails at once
            if st.button("⚡ Generate All Shortlist Emails at Once",
                         key="gen_all_shortlist", use_container_width=False):
                with st.spinner("Generating shortlist emails..."):
                    for res in qualified_for_email:
                        cand_key = f"shortlist_{res.get('name','?')}_{res.get('upload_pos',0)}"
                        draft = generate_email_draft(
                            candidate=res,
                            job_description=_email_jd,
                            school_name=school_name,
                            hr_name=hr_name,
                            groq_api_key=groq_key,
                            groq_model=groq_model,
                        )
                        st.session_state.email_drafts[cand_key] = draft
                st.success(f"✅ Generated {len(qualified_for_email)} shortlist email(s)!")
                st.rerun()

            st.markdown("<br>", unsafe_allow_html=True)
            for res in qualified_for_email:
                rank_label = f"Rank #{res.get('rank')} · " if res.get("rank") else ""
                score_label = f"{res.get('score', 0):.0f}% match"
                st.markdown(
                    f'<p style="font-size:1rem;font-weight:700;color:{C["sage"]};margin:8px 0 6px;">'
                    f'✅ {res.get("name","?")} &nbsp;<span style="font-size:.8rem;'
                    f'color:{C["silver"]};font-weight:400;">{rank_label}{score_label}</span></p>',
                    unsafe_allow_html=True
                )
                _render_email_card(res, "shortlist")

    # ── Rejection emails tab ──────────────────────────────────────────────
    with email_tab_rejection:
        if not filtered_for_email:
            st.success("🎉 No candidates were filtered out — no rejection emails needed!")
        else:
            st.markdown(
                f'<p style="font-size:.83rem;color:{C["silver"]};margin:0 0 16px;">'
                f'Generate polite, empathetic rejection emails with specific reasons for each filtered candidate.</p>',
                unsafe_allow_html=True
            )
            if st.button("⚡ Generate All Rejection Emails at Once",
                         key="gen_all_rejection", use_container_width=False):
                with st.spinner("Generating rejection emails..."):
                    for res in filtered_for_email:
                        cand_key = f"rejection_{res.get('name','?')}_{res.get('upload_pos',0)}"
                        draft = generate_email_draft(
                            candidate=res,
                            job_description=_email_jd,
                            school_name=school_name,
                            hr_name=hr_name,
                            groq_api_key=groq_key,
                            groq_model=groq_model,
                        )
                        st.session_state.email_drafts[cand_key] = draft
                st.success(f"✅ Generated {len(filtered_for_email)} rejection email(s)!")
                st.rerun()

            st.markdown("<br>", unsafe_allow_html=True)
            for res in filtered_for_email:
                score_label = f"{res.get('score', 0):.0f}% match"
                reasons = "; ".join(res.get("fail_reasons", []))[:80]
                st.markdown(
                    f'<p style="font-size:1rem;font-weight:700;color:{C["rose"]};margin:8px 0 4px;">'
                    f'❌ {res.get("name","?")} &nbsp;<span style="font-size:.78rem;'
                    f'color:{C["silver"]};font-weight:400;">{score_label}</span></p>'
                    f'<p style="font-size:.75rem;color:{C["silver"]};margin:0 0 8px;">{reasons}</p>',
                    unsafe_allow_html=True
                )
                _render_email_card(res, "rejection")

    # ── Bulk download tab ─────────────────────────────────────────────────
    with email_tab_bulk:
        st.markdown(
            f'<p style="font-size:.83rem;color:{C["silver"]};margin:0 0 16px;">'
            f'Generate and download all emails (shortlist + rejection) as a single .txt file.</p>',
            unsafe_allow_html=True
        )

        if st.button("⚡ Generate ALL Emails (Shortlist + Rejection)",
                     key="gen_all_bulk", use_container_width=False):
            with st.spinner("Generating all emails..."):
                all_emails = generate_all_emails(
                    results=_email_results,
                    job_description=_email_jd,
                    school_name=school_name,
                    hr_name=hr_name,
                    groq_api_key=groq_key,
                    groq_model=groq_model,
                )
                st.session_state["bulk_emails"] = all_emails
            st.success(f"✅ Generated {len(all_emails)} email(s) total!")
            st.rerun()

        if "bulk_emails" in st.session_state and st.session_state.bulk_emails:
            bulk = st.session_state.bulk_emails
            # Build combined text file
            lines = ["=" * 60, "EduHire — Email Drafts Export", "=" * 60, ""]
            shortlisted = [e for e in bulk if e["type"] == "shortlist"]
            rejected    = [e for e in bulk if e["type"] == "rejection"]

            lines += [f"SHORTLISTED CANDIDATES ({len(shortlisted)})", "-" * 40, ""]
            for e in shortlisted:
                lines += [
                    f"To: {e['name']}",
                    f"Subject: {e['subject']}",
                    "",
                    e["body"],
                    "",
                    "=" * 60,
                    "",
                ]

            lines += [f"REJECTED CANDIDATES ({len(rejected)})", "-" * 40, ""]
            for e in rejected:
                lines += [
                    f"To: {e['name']}",
                    f"Subject: {e['subject']}",
                    "",
                    e["body"],
                    "",
                    "=" * 60,
                    "",
                ]

            combined_text = "\n".join(lines)

            st.download_button(
                label="⬇️  Download All Emails (.txt)",
                data=combined_text,
                file_name="eduhire_all_email_drafts.txt",
                mime="text/plain",
                key="dl_bulk_all",
            )

            # Summary table
            st.markdown(
                f'<p style="font-size:.9rem;font-weight:700;color:{C["cream"]};margin:20px 0 8px;">📋 Generated Emails Summary</p>',
                unsafe_allow_html=True
            )
            for e in bulk:
                icon  = "✅" if e["type"] == "shortlist" else "❌"
                score = f"{e.get('score', 0):.0f}%"
                st.markdown(
                    f'<div style="display:flex;align-items:center;justify-content:space-between;'
                    f'padding:8px 14px;background:rgba(13,33,64,.7);border-radius:10px;'
                    f'margin-bottom:6px;border:1px solid rgba(30,58,95,.5);">'
                    f'<span style="font-size:.88rem;color:{C["cream"]};">{icon} {e["name"]}</span>'
                    f'<span style="font-size:.78rem;color:{C["silver"]};">{e["subject"][:55]}...</span>'
                    f'<span style="font-size:.78rem;color:{C["sky"]};">{score}</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )

else:
    st.markdown(
        f'<div style="text-align:center;padding:32px;color:{C["silver"]};'
        f'font-size:.88rem;border:1px dashed rgba(148,163,184,.3);'
        f'border-radius:12px;margin-bottom:12px;">'
        f'📧 Run the screening first, then come back here to generate emails for all candidates.</div>',
        unsafe_allow_html=True
    )


# ══════════════════════════════════════════════════════════════════════════════
# GOOGLE CALENDAR SECTION — Auto-Schedule Interview Slots
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(divider(), unsafe_allow_html=True)

st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(16,185,129,.12),rgba(14,165,233,.12));
     border:1px solid rgba(16,185,129,.3);border-radius:16px;
     padding:28px 36px;margin-bottom:24px;">
  <div style="display:flex;align-items:center;gap:16px;margin-bottom:10px;">
    <div style="font-size:2rem;">🗓️</div>
    <div>
      <h2 style="margin:0;font-size:1.55rem;font-family:'DM Serif Display',Georgia,serif;
                 color:{C['cream']};">Google Calendar — Auto-Schedule Interviews</h2>
      <p style="margin:4px 0 0;color:{C['silver']};font-size:.88rem;">
        Instantly book interview slots for all shortlisted candidates on your Google Calendar,
        with Google Meet links and email invites sent automatically.
      </p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if not (_CALENDAR_MODULE and _GOOGLE_AVAILABLE):
    st.warning(
        "🔌 Google Calendar integration is not available. "
        "Install the required packages and ensure calendar_scheduler.py is present:\n\n"
        "`pip install google-auth google-auth-oauthlib google-api-python-client`",
        icon="📦",
    )
elif "screening_results" not in st.session_state or not st.session_state.screening_results:
    st.info(
        "📋 Run the resume screening first. Once shortlisted candidates appear, "
        "you can auto-schedule their interviews here.",
        icon="ℹ️",
    )
else:
    _cal_results   = st.session_state.screening_results
    _cal_jd        = st.session_state.get("screening_jd", "")
    _qualified_cal = [r for r in _cal_results if not r["filtered"]]

    if not _qualified_cal:
        st.info("No shortlisted candidates to schedule. Adjust your screening thresholds.", icon="ℹ️")
    else:
        st.markdown(
            f'<p style="font-size:.9rem;color:{C["silver"]};margin-bottom:16px;">' +
            f'Found <b style="color:{C["sage"]};">{len(_qualified_cal)} shortlisted candidate(s)</b>. ' +
            f'Select the ones you want to schedule below.</p>',
            unsafe_allow_html=True
        )

        # ── STEP 1: Candidate Selection ──────────────────────────────────────
        st.markdown(
            f'<p style="font-size:1rem;font-weight:700;color:{C["cream"]};margin:0 0 10px;">👥 Step 1 — Select Candidates to Schedule</p>',
            unsafe_allow_html=True
        )

        # Init selection state
        if "cal_selected_candidates" not in st.session_state:
            st.session_state.cal_selected_candidates = set()

        # Select All / Clear All buttons
        sel_btn_col1, sel_btn_col2, sel_btn_col3 = st.columns([1, 1, 4])
        with sel_btn_col1:
            if st.button("✅ Select All", key="cal_sel_all", use_container_width=True):
                st.session_state.cal_selected_candidates = {c["name"] for c in _qualified_cal}
                st.rerun()
        with sel_btn_col2:
            if st.button("⬜ Clear All", key="cal_clear_all", use_container_width=True):
                st.session_state.cal_selected_candidates = set()
                st.rerun()

        # Per-candidate checkboxes in a grid
        chk_cols = st.columns(min(len(_qualified_cal), 3))
        for ci, cand in enumerate(_qualified_cal):
            cname = cand["name"]
            score = cand.get("score", 0)
            rank  = cand.get("rank", ci + 1)
            with chk_cols[ci % 3]:
                checked = cname in st.session_state.cal_selected_candidates
                new_val = st.checkbox(
                    f"#{rank} · {cname}  ({score:.1f}%)",
                    value=checked,
                    key=f"cal_chk_{ci}",
                )
                if new_val:
                    st.session_state.cal_selected_candidates.add(cname)
                elif cname in st.session_state.cal_selected_candidates:
                    st.session_state.cal_selected_candidates.discard(cname)

        selected_names = st.session_state.cal_selected_candidates
        selected_cands = [c for c in _qualified_cal if c["name"] in selected_names]

        if not selected_cands:
            st.warning("⚠️ Please select at least one candidate to schedule.", icon="👆")
        else:
            st.markdown(
                f'<p style="font-size:.82rem;color:{C["sage"]};margin:8px 0 16px;">' +
                f'✔ {len(selected_cands)} candidate(s) selected for scheduling.</p>',
                unsafe_allow_html=True
            )

        st.markdown("<hr style='border-color:rgba(255,255,255,.1);margin:20px 0;'>", unsafe_allow_html=True)

        # ── STEP 2: Scheduling Settings ──────────────────────────────────────
        st.markdown(
            f'<p style="font-size:1rem;font-weight:700;color:{C["cream"]};margin:0 0 12px;">⚙️ Step 2 — Scheduling Settings</p>',
            unsafe_allow_html=True
        )

        cal_col1, cal_col2, cal_col3 = st.columns(3)

        import datetime as _dt

        with cal_col1:
            cal_start_date = st.date_input(
                "📅 Start scheduling from",
                value=_dt.date.today() + _dt.timedelta(days=1),
                min_value=_dt.date.today(),
                help="First possible date for interviews."
            )
            cal_duration = st.selectbox(
                "⏱ Interview Duration",
                options=[30, 45, 60, 90],
                index=1,
                format_func=lambda x: f"{x} minutes",
            )

        with cal_col2:
            cal_start_hour = st.slider("🕘 Working Hours Start", 7, 12, 9)
            cal_end_hour   = st.slider("🕔 Working Hours End",  13, 20, 17)
            cal_gap        = st.selectbox(
                "⏸ Gap Between Slots",
                options=[10, 15, 20, 30],
                index=1,
                format_func=lambda x: f"{x} min break",
            )

        with cal_col3:
            cal_skip_weekends = st.checkbox("⛔ Skip Weekends", value=True)
            cal_add_meet      = st.checkbox("📹 Add Google Meet Link", value=True)
            cal_check_avail   = st.checkbox("🔍 Check Calendar Availability", value=True,
                                             help="Uses freebusy API to skip already-booked slots.")

        st.markdown("<hr style='border-color:rgba(255,255,255,.1);margin:20px 0;'>", unsafe_allow_html=True)

        # ── STEP 3: Candidate Emails ─────────────────────────────────────────
        st.markdown(
            f'<p style="font-size:1rem;font-weight:700;color:{C["cream"]};margin:0 0 6px;">📧 Step 3 — Candidate Emails (optional)</p>',
            unsafe_allow_html=True
        )
        st.markdown(
            f'<p style="font-size:.82rem;color:{C["silver"]};margin-bottom:12px;">' +
            f'Only shown for selected candidates. Leave blank to create events without sending invites.</p>',
            unsafe_allow_html=True
        )

        if "cal_candidate_emails" not in st.session_state:
            st.session_state.cal_candidate_emails = {}

        if selected_cands:
            email_cols = st.columns(min(len(selected_cands), 3))
            for ci, cand in enumerate(selected_cands):
                cname = cand["name"]
                with email_cols[ci % 3]:
                    em = st.text_input(
                        f"📧 {cname}",
                        value=st.session_state.cal_candidate_emails.get(cname, ""),
                        placeholder="candidate@email.com",
                        key=f"cal_email_{ci}",
                    )
                    st.session_state.cal_candidate_emails[cname] = em
        else:
            st.markdown(
                f'<p style="font-size:.82rem;color:{C["silver"]};">Select candidates above to enter their emails.</p>',
                unsafe_allow_html=True
            )

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Schedule Button ──────────────────────────────────────────────────
        cal_btn_col1, cal_btn_col2 = st.columns([2, 3])

        with cal_btn_col1:
            n_sel = len(selected_cands)
            do_schedule = st.button(
                f"🗓️ Schedule {n_sel} Interview{'s' if n_sel != 1 else ''}",
                use_container_width=True,
                key="cal_schedule_btn",
                disabled=(n_sel == 0),
                help="Creates Google Calendar events only for the candidates you checked above.",
            )

        with cal_btn_col2:
            st.markdown(
                f'<p style="font-size:.78rem;color:{C["silver"]};padding-top:10px;">' +
                f'Opens a browser tab for Google sign-in on first use. ' +
                f'Saved token is reused on subsequent runs.</p>',
                unsafe_allow_html=True
            )

        if do_schedule:
            if not selected_cands:
                st.error("⚠️ No candidates selected. Please check at least one candidate above.", icon="👆")
            elif not gcal_interviewer_email:
                st.error("⚠️ Please enter the Interviewer Email in the sidebar before scheduling.", icon="📧")
            else:
                with st.spinner("🔐 Authenticating with Google Calendar..."):
                    scheduler = CalendarScheduler(
                        credentials_json_path=gcal_creds_path,
                        calendar_id=gcal_calendar_id,
                    )
                    auth_ok = scheduler.authenticate()

                if not auth_ok:
                    st.error(f"❌ Authentication failed: {scheduler.error}", icon="🔐")
                else:
                    st.success(f"✅ Authenticated! Scheduling {len(selected_cands)} interview(s)...", icon="🔐")

                    # Build ONLY selected candidates list — this is what drives slot count
                    candidates_to_schedule = [
                        {
                            "name":  cand["name"],
                            "email": st.session_state.cal_candidate_emails.get(cand["name"], ""),
                        }
                        for cand in selected_cands   # ← only selected, not all shortlisted
                    ]

                    schedule_config = {
                        "start_date":         cal_start_date,
                        "duration_minutes":   cal_duration,
                        "working_hours":      (cal_start_hour, cal_end_hour),
                        "skip_weekends":      cal_skip_weekends,
                        "gap_minutes":        cal_gap,
                        "interviewer_name":   hr_name,
                        "interviewer_email":  gcal_interviewer_email,
                        "school_name":        school_name,
                        "job_title":          "Teacher Position",
                        "add_meet":           cal_add_meet,
                        "check_availability": cal_check_avail,
                        "timezone":           "Asia/Kolkata",
                    }

                    with st.spinner(f"📅 Creating {len(candidates_to_schedule)} calendar event(s)..."):
                        schedule_results = scheduler.schedule_interviews(
                            candidates=candidates_to_schedule,   # exact selected list
                            config=schedule_config,
                        )

                    st.session_state["cal_schedule_results"] = schedule_results

        # ── Display scheduling results ────────────────────────────────────────
        if "cal_schedule_results" in st.session_state:
            sched_results = st.session_state["cal_schedule_results"]
            success_count = sum(1 for r in sched_results if r.get("success"))
            fail_count    = len(sched_results) - success_count

            st.markdown(
                f'<div style="background:rgba(16,185,129,.1);border:1px solid rgba(16,185,129,.3);' +
                f'border-radius:12px;padding:16px 20px;margin:16px 0;">' +
                f'<p style="margin:0;font-weight:700;color:{C["sage"]};">📊 Scheduling Summary</p>' +
                f'<p style="margin:4px 0 0;color:{C["cream"]};font-size:.88rem;">' +
                f'Scheduled for: <b>{len(sched_results)}</b> selected candidate(s) &nbsp;·&nbsp; ' +
                f'✅ {success_count} succeeded &nbsp;·&nbsp; ' +
                f'{"❌ " + str(fail_count) + " failed" if fail_count else "All successful 🎉"}</p>' +
                f'</div>',
                unsafe_allow_html=True
            )

            for res in sched_results:
                is_ok = res.get("success", False)
                bg    = "rgba(16,185,129,.08)" if is_ok else "rgba(244,63,94,.08)"
                border= "rgba(16,185,129,.3)"  if is_ok else "rgba(244,63,94,.3)"
                icon  = "✅" if is_ok else "❌"

                start_str = res["start"].strftime("%a %d %b %Y · %I:%M %p") if res.get("start") else "—"
                end_str   = res["end"].strftime("%I:%M %p")                   if res.get("end")   else "—"
                meet_link = res.get("meet_link",  "")
                ev_link   = res.get("event_link", "")

                meet_html = (
                    f'<a href="{meet_link}" target="_blank" style="color:{C["sky"]};font-weight:700;">' +
                    f'📹 Join Meet</a> &nbsp;·&nbsp; '
                    if meet_link else ""
                )
                cal_html = (
                    f'<a href="{ev_link}" target="_blank" style="color:{C["teal"]};font-weight:700;">' +
                    f'📅 Open in Calendar</a>'
                    if ev_link else ""
                )
                err_html = (
                    f'<p style="color:#FDA4AF;font-size:.82rem;margin:4px 0 0;">⚠️ {res.get("error","")}</p>'
                    if not is_ok else ""
                )

                st.markdown(
                    f'<div style="background:{bg};border:1px solid {border};' +
                    f'border-radius:10px;padding:14px 18px;margin-bottom:10px;">' +
                    f'<p style="margin:0;font-weight:700;color:{C["cream"]};font-size:.95rem;">' +
                    f'{icon} {res["name"]}</p>' +
                    f'<p style="margin:4px 0 0;color:{C["silver"]};font-size:.82rem;">' +
                    f'🕐 {start_str} — {end_str}</p>' +
                    (f'<p style="margin:6px 0 0;font-size:.82rem;">{meet_html}{cal_html}</p>' if is_ok else "") +
                    err_html +
                    f'</div>',
                    unsafe_allow_html=True
                )

            # ── Download summary as CSV ───────────────────────────────────────
            import csv, io as _io
            csv_buf = _io.StringIO()
            writer  = csv.writer(csv_buf)
            writer.writerow(["Candidate", "Email", "Date & Time", "Meet Link", "Calendar Link", "Status", "Error"])
            for res in sched_results:
                writer.writerow([
                    res.get("name",  ""),
                    res.get("email", ""),
                    res["start"].strftime("%Y-%m-%d %H:%M") if res.get("start") else "",
                    res.get("meet_link",  ""),
                    res.get("event_link", ""),
                    "Scheduled" if res.get("success") else "Failed",
                    res.get("error", ""),
                ])
            st.download_button(
                label="⬇️ Download Schedule CSV",
                data=csv_buf.getvalue(),
                file_name="interview_schedule.csv",
                mime="text/csv",
                key="cal_download_csv",
            )

# ══════════════════════════════════════════════════════════════════════════════
# PRE-INTERVIEW BRIEF SECTION  (Medium Impact Feature #6)
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(divider(), unsafe_allow_html=True)

st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(245,158,11,.10),rgba(15,118,110,.10));
     border:1px solid rgba(245,158,11,.30);border-radius:16px;
     padding:28px 36px;margin-bottom:24px;">
  <div style="display:flex;align-items:center;gap:16px;margin-bottom:10px;">
    <div style="font-size:2rem;">📋</div>
    <div>
      <h2 style="font-size:1.4rem;font-weight:800;color:{C['cream']};margin:0;
        font-family:'DM Serif Display',Georgia,serif;">Pre-Interview Brief Generator</h2>
      <p style="margin:4px 0 0;color:{C['silver']};font-size:.88rem;">
        Auto-compiles candidate background, last call notes, and open questions
        before each scheduled interview — ready to print or download.
      </p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if not _BRIEF_MODULE:
    st.warning("⚠️ interview_brief.py not found. Please ensure it is in the project folder.", icon="📄")
elif "screening_results" not in st.session_state or not st.session_state.screening_results:
    st.info("📋 Run the resume screening first to generate pre-interview briefs.", icon="ℹ️")
else:
    _brief_results  = st.session_state.screening_results
    _brief_pipeline = st.session_state.get("pipeline", {})
    _brief_jd       = st.session_state.get("screening_jd", "")
    _qual_brief     = [r for r in _brief_results if not r.get("filtered")]

    if not _qual_brief:
        st.info("No shortlisted candidates available. Adjust your screening thresholds.", icon="ℹ️")
    else:
        # ── Candidate selector ────────────────────────────────────────────
        _brief_names = [r.get("name", f"Candidate {i+1}") for i, r in enumerate(_qual_brief)]
        _brief_sel_name = st.selectbox(
            "Select candidate to generate brief for:",
            options=_brief_names,
            key="brief_candidate_select",
        )
        _brief_cand = next((r for r in _qual_brief if r.get("name") == _brief_sel_name), _qual_brief[0])

        # Pipeline data for this candidate
        _brief_pipe_key  = f"{_brief_cand.get('name','unknown')}_{_brief_cand.get('upload_pos',0)}"
        _brief_pipe_data = _brief_pipeline.get(_brief_pipe_key, {})

        # ── Interview time (from calendar results if available) ───────────
        _brief_interview_dt = None
        if "cal_schedule_results" in st.session_state:
            for sr in st.session_state["cal_schedule_results"]:
                if sr.get("name") == _brief_cand.get("name") and sr.get("success"):
                    _brief_interview_dt = sr.get("start")
                    break

        col_brief_left, col_brief_right = st.columns([2, 1])

        with col_brief_right:
            st.markdown(
                f'<p style="font-size:.9rem;font-weight:700;color:{C["cream"]};margin-bottom:8px;">⚙️ Brief Options</p>',
                unsafe_allow_html=True,
            )
            _brief_interviewer = st.text_input(
                "Interviewer name", value=school_name_val if "school_name_val" in dir() else "Hiring Manager",
                key="brief_interviewer_name",
            )
            _brief_school = st.text_input(
                "School name", value=school_name_val if "school_name_val" in dir() else "Our School",
                key="brief_school_name_input",
            )
            _brief_job = st.text_input(
                "Job title", value="Teacher Position", key="brief_job_title",
            )
            _brief_use_ai = st.toggle(
                "✨ Add AI narrative summary",
                value=False,
                key="brief_use_ai",
                help="Requires Groq API key (configured in sidebar)",
            )
            _brief_extra_q_raw = st.text_area(
                "Add your own questions (one per line):",
                height=100,
                key="brief_extra_questions",
                placeholder="e.g. How do you handle classroom management?\nDescribe a successful parent-teacher interaction.",
            )
            _brief_extra_qs = [q.strip() for q in _brief_extra_q_raw.strip().splitlines() if q.strip()]

        with col_brief_left:
            if st.button("📋 Generate Brief", key="gen_brief_btn", type="primary", use_container_width=True):
                with st.spinner("Assembling pre-interview brief…"):
                    _brief_data = generate_brief(
                        candidate        = _brief_cand,
                        pipeline_data    = _brief_pipe_data,
                        interview_dt     = _brief_interview_dt,
                        school_name      = _brief_school,
                        interviewer_name = _brief_interviewer,
                        job_title        = _brief_job,
                        extra_questions  = _brief_extra_qs or None,
                    )

                    _brief_ai_summary = ""
                    if _brief_use_ai:
                        _groq_key_for_brief = groq_key or ""
                        if _groq_key_for_brief:
                            _brief_ai_summary = enrich_brief_with_ai(_brief_data, _groq_key_for_brief, groq_model)
                        else:
                            st.warning("Add a Groq API key in the sidebar to enable AI narrative summaries.", icon="🔑")

                    st.session_state["current_brief"]           = _brief_data
                    st.session_state["current_brief_ai"]        = _brief_ai_summary
                    st.session_state["current_brief_cand_name"] = _brief_sel_name

        # ── Render brief if generated ────────────────────────────────────
        if "current_brief" in st.session_state and st.session_state.get("current_brief_cand_name") == _brief_sel_name:
            _bd    = st.session_state["current_brief"]
            _bai   = st.session_state.get("current_brief_ai", "")
            _bbg   = _bd["background"]
            _bsc   = _bd["scores"]
            _bsk   = _bd["skills"]
            _bnotes= _bd["last_notes"]
            _bhist = _bd["stage_history"]
            _bqs   = _bd["open_questions"]
            _bgen  = _brief_fmt_ts(_bd.get("generated_at",""))

            # Header card
            import datetime as _bdt_mod
            _bdt_str = (
                _bbg["interview_dt"].strftime("%d %b %Y at %I:%M %p")
                if isinstance(_bbg.get("interview_dt"), _bdt_mod.datetime)
                else ("Not yet scheduled")
            )
            st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(13,27,42,.95),rgba(15,52,96,.80));
     border:1.5px solid rgba(245,158,11,.35);border-radius:14px;
     padding:22px 28px;margin:18px 0 10px;">
  <div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;gap:12px;">
    <div>
      <p style="margin:0;font-size:.75rem;font-weight:700;color:{C['gold']};text-transform:uppercase;letter-spacing:2px;">Pre-Interview Brief</p>
      <h3 style="margin:4px 0 2px;font-size:1.6rem;font-weight:900;color:{C['cream']};">{_bbg['name']}</h3>
      <p style="margin:0;color:{C['silver']};font-size:.88rem;">{_bbg['job_title']} &nbsp;·&nbsp; {_bbg['school']}</p>
    </div>
    <div style="text-align:right;">
      <p style="margin:0;font-size:.78rem;color:{C['gold']};font-weight:700;">🗓 {_bdt_str}</p>
      <p style="margin:4px 0 0;font-size:.78rem;color:{C['silver']};">Interviewer: <b style="color:{C['cream']};">{_bbg['interviewer']}</b></p>
      <p style="margin:4px 0 0;font-size:.75rem;color:{C['gray']};">Generated {_bgen}</p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

            # Tabs for brief sections
            _btab1, _btab2, _btab3, _btab4 = st.tabs(["📊 Scores & Skills", "📝 Call Notes", "❓ Open Questions", "✨ AI Summary"])

            with _btab1:
                _bsc_col1, _bsc_col2 = st.columns(2)
                with _bsc_col1:
                    st.markdown(f"""
<div style="background:rgba(13,27,42,.7);border:1px solid rgba(30,58,95,.8);border-radius:12px;padding:16px 20px;">
  <p style="margin:0 0 12px;font-size:.78rem;font-weight:800;color:{C['gold']};text-transform:uppercase;letter-spacing:1.5px;">📈 ATS Scores</p>
  <table style="width:100%;border-collapse:collapse;font-size:.87rem;">
    <tr><td style="color:{C['silver']};padding:4px 0;">Overall</td>
        <td style="text-align:right;font-weight:800;color:{'#10B981' if _bsc['overall']>=70 else ('#F59E0B' if _bsc['overall']>=50 else '#F43F5E')};font-size:1rem;">{_bsc['overall']}/100</td></tr>
    <tr><td style="color:{C['silver']};padding:4px 0;">Rank</td>
        <td style="text-align:right;color:{C['cream']};font-weight:700;">#{_bsc.get('rank','N/A')}</td></tr>
    <tr><td style="color:{C['silver']};padding:4px 0;">Keyword Match</td>
        <td style="text-align:right;color:{C['cream']};">{_bsc['keyword_match']}%</td></tr>
    <tr><td style="color:{C['silver']};padding:4px 0;">Experience</td>
        <td style="text-align:right;color:{C['cream']};">{_bsc['experience']}%</td></tr>
    <tr><td style="color:{C['silver']};padding:4px 0;">Education</td>
        <td style="text-align:right;color:{C['cream']};">{_bsc['education']}%</td></tr>
    <tr><td style="color:{C['silver']};padding:4px 0;">Certifications</td>
        <td style="text-align:right;color:{C['cream']};">{_bsc['certifications']}%</td></tr>
  </table>
</div>
""", unsafe_allow_html=True)

                with _bsc_col2:
                    _matched_html = "".join(
                        f'<span style="background:rgba(16,185,129,.15);border:1px solid rgba(16,185,129,.4);'
                        f'border-radius:6px;padding:3px 8px;font-size:.78rem;color:#6EE7B7;margin:2px;">{kw}</span>'
                        for kw in _bsk["matched"][:12]
                    ) or f'<span style="color:{C["silver"]};">—</span>'
                    _missing_html = "".join(
                        f'<span style="background:rgba(244,63,94,.12);border:1px solid rgba(244,63,94,.35);'
                        f'border-radius:6px;padding:3px 8px;font-size:.78rem;color:#FDA4AF;margin:2px;">{kw}</span>'
                        for kw in _bsk["missing"][:10]
                    ) or f'<span style="color:{C["silver"]};">—</span>'
                    st.markdown(f"""
<div style="background:rgba(13,27,42,.7);border:1px solid rgba(30,58,95,.8);border-radius:12px;padding:16px 20px;height:100%;">
  <p style="margin:0 0 8px;font-size:.78rem;font-weight:800;color:{C['sage']};text-transform:uppercase;letter-spacing:1.5px;">✅ Matched Skills</p>
  <div style="margin-bottom:14px;display:flex;flex-wrap:wrap;gap:4px;">{_matched_html}</div>
  <p style="margin:0 0 8px;font-size:.78rem;font-weight:800;color:{C['rose']};text-transform:uppercase;letter-spacing:1.5px;">⚠️ Missing Skills</p>
  <div style="display:flex;flex-wrap:wrap;gap:4px;">{_missing_html}</div>
</div>
""", unsafe_allow_html=True)

            with _btab2:
                if not _bnotes and not _bhist:
                    st.info("No call notes or stage history logged yet for this candidate.", icon="📭")
                else:
                    if _bnotes:
                        st.markdown(
                            f'<p style="font-size:.85rem;font-weight:700;color:{C["gold"]};margin-bottom:8px;">🗒 Last Call / Meeting Notes</p>',
                            unsafe_allow_html=True,
                        )
                        for _n in _bnotes:
                            st.markdown(f"""
<div style="background:rgba(13,27,42,.6);border-left:3px solid rgba(245,158,11,.6);
     border-radius:0 10px 10px 0;padding:10px 16px;margin-bottom:8px;">
  <p style="margin:0 0 4px;font-size:.75rem;color:{C['gray']};">
    <b style="color:{C['silver']};">{_n.get('author','Recruiter')}</b> &nbsp;·&nbsp; {_brief_fmt_ts(_n.get('ts',''))}
  </p>
  <p style="margin:0;font-size:.88rem;color:{C['cream']};">{_n.get('text','')}</p>
</div>
""", unsafe_allow_html=True)

                    if _bhist:
                        st.markdown(
                            f'<p style="font-size:.85rem;font-weight:700;color:{C["sky"]};margin:14px 0 8px;">🔄 Pipeline Stage History</p>',
                            unsafe_allow_html=True,
                        )
                        for _h in _bhist:
                            _hstage = _h.get("stage","?")
                            _hcolor = STAGE_COLORS.get(_hstage, ("#1E3A5F","#94A3B8"))
                            st.markdown(f"""
<div style="display:flex;align-items:center;gap:12px;margin-bottom:6px;">
  <span style="background:{_hcolor[0]};color:{_hcolor[1]};border-radius:8px;
               padding:3px 10px;font-size:.78rem;font-weight:700;white-space:nowrap;">
    {STAGE_ICONS.get(_hstage,'📋')} {_hstage}
  </span>
  <span style="font-size:.78rem;color:{C['silver']};">{_brief_fmt_ts(_h.get('ts',''))}</span>
  <span style="font-size:.82rem;color:{C['gray']};font-style:italic;">{_h.get('note','')}</span>
</div>
""", unsafe_allow_html=True)

            with _btab3:
                st.markdown(
                    f'<p style="font-size:.85rem;color:{C["silver"]};margin-bottom:12px;">'
                    f'These questions were auto-generated based on the candidate\'s profile and gaps. '
                    f'Add your own via the options panel, then download the brief.</p>',
                    unsafe_allow_html=True,
                )
                for _qi, _q in enumerate(_bqs, 1):
                    st.markdown(f"""
<div style="background:rgba(14,165,233,.08);border:1px solid rgba(14,165,233,.25);
     border-radius:10px;padding:12px 16px;margin-bottom:8px;display:flex;gap:12px;align-items:flex-start;">
  <span style="font-size:1rem;font-weight:800;color:{C['sky']};min-width:22px;">Q{_qi}</span>
  <p style="margin:0;font-size:.88rem;color:{C['cream']};line-height:1.6;">{_q}</p>
</div>
""", unsafe_allow_html=True)

            with _btab4:
                if _bai:
                    st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(139,92,246,.10),rgba(14,165,233,.08));
     border:1px solid rgba(139,92,246,.30);border-radius:12px;padding:20px 24px;">
  <p style="margin:0 0 10px;font-size:.78rem;font-weight:800;color:#C4B5FD;text-transform:uppercase;letter-spacing:1.5px;">✨ AI Narrative Summary</p>
  <p style="margin:0;font-size:.9rem;color:{C['cream']};line-height:1.75;">{_bai.replace(chr(10),'<br>')}</p>
</div>
""", unsafe_allow_html=True)
                else:
                    if _brief_use_ai and not groq_key:
                        st.warning("Add a Groq API key in the sidebar to enable AI narrative summaries.", icon="🔑")
                    elif not _brief_use_ai:
                        st.info("Enable '✨ Add AI narrative summary' in the options panel and regenerate.", icon="💡")

            # ── Download button ───────────────────────────────────────────
            st.markdown("<div style='margin-top:16px;'></div>", unsafe_allow_html=True)
            _brief_text_output = brief_to_text(_bd)
            if _bai:
                _brief_text_output += f"\n── AI NARRATIVE SUMMARY ──────────────────────────────────────\n{_bai}\n{'='*66}\n"
            _brief_filename = f"brief_{_bbg['name'].replace(' ','_').lower()}_{__import__('datetime').datetime.now().strftime('%Y%m%d')}.txt"
            st.download_button(
                label="⬇️ Download Brief (.txt)",
                data=_brief_text_output.encode("utf-8"),
                file_name=_brief_filename,
                mime="text/plain",
                key="brief_download_btn",
            )

# ══════════════════════════════════════════════════════════════════════════════
# CHATBOT SECTION — AI Hiring Assistant (EduBot)
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(divider(), unsafe_allow_html=True)

st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(14,165,233,.12),rgba(15,118,110,.12));
     border:1px solid rgba(14,165,233,.25);border-radius:16px;
     padding:28px 36px;margin-bottom:24px;">
  <div style="display:flex;align-items:center;gap:16px;margin-bottom:10px;">
    <div style="font-size:2rem;">🤖</div>
    <div>
      <h2 style="font-size:1.5rem;font-weight:800;color:#E8EFF8;margin:0;
        font-family:'DM Serif Display',Georgia,serif;">EduBot — AI Hiring Assistant</h2>
      <p style="font-size:.83rem;color:{C['silver']};margin:4px 0 0;">
        Ask anything about the screened candidates, scores, keyword gaps, or hiring recommendations.
      </p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── Session state init ────────────────────────────────────────────────────────
if "chat_history" not in st.session_state:
    st.session_state.chat_history = []
if "chat_results_snapshot" not in st.session_state:
    st.session_state.chat_results_snapshot = []
if "chat_jd_snapshot" not in st.session_state:
    st.session_state.chat_jd_snapshot = ""
if "rag_index" not in st.session_state:
    st.session_state.rag_index = None

if "screening_results" in st.session_state:
    st.session_state.chat_results_snapshot = st.session_state.screening_results
if "screening_jd" in st.session_state:
    st.session_state.chat_jd_snapshot = st.session_state.screening_jd

# ── OpenAI status banner ─────────────────────────────────────────────────────
if groq_key:
    st.markdown(
        f'<div style="background:rgba(16,185,129,.12);border:1px solid rgba(16,185,129,.35);'
        f'border-radius:10px;padding:10px 16px;margin-bottom:12px;font-size:.83rem;">'
        f'✅ <strong>OpenAI GPT active</strong> — Ask me <em>anything</em>: candidate analysis, '
        f'hiring strategy, salary benchmarks, or any general question!</div>',
        unsafe_allow_html=True
    )
else:
    st.markdown(
        f'<div style="background:rgba(251,191,36,.10);border:1px solid rgba(251,191,36,.3);'
        f'border-radius:10px;padding:10px 16px;margin-bottom:12px;font-size:.83rem;">'
        f'⚠️ <strong>No Groq key</strong> — Running in rule-based mode. '
        f'Add your <code>sk-...</code> key in the sidebar to unlock full AI chat.</div>',
        unsafe_allow_html=True
    )

# ── Suggested questions ───────────────────────────────────────────────────────
suggestions = get_suggested_questions(st.session_state.chat_results_snapshot)
st.markdown(
    f'<p style="font-size:.82rem;font-weight:600;color:{C["sky"]};margin:0 0 8px;">💡 Quick Questions</p>',
    unsafe_allow_html=True
)
sug_cols = st.columns(len(suggestions))
for i, (col, question) in enumerate(zip(sug_cols, suggestions)):
    with col:
        if st.button(question, key=f"sug_{i}", use_container_width=True):
            st.session_state.chat_history.append({"role": "user", "content": question})
            with st.spinner("EduBot is thinking..."):
                reply = get_chatbot_response(
                    user_message=question,
                    chat_history=st.session_state.chat_history[:-1],
                    screening_results=st.session_state.chat_results_snapshot,
                    job_description=st.session_state.chat_jd_snapshot,
                    groq_api_key=groq_key,
                    groq_model=groq_model,
                    rag_index=st.session_state.get("rag_index"),
                )
            st.session_state.chat_history.append({"role": "assistant", "content": reply})
            st.rerun()

st.markdown("<br>", unsafe_allow_html=True)

# ── Chat history display ──────────────────────────────────────────────────────
with st.container():
    if not st.session_state.chat_history:
        st.markdown(
            f'<div style="text-align:center;padding:32px;color:{C["silver"]};'
            f'font-size:.88rem;border:1px dashed rgba(148,163,184,.3);'
            f'border-radius:12px;margin-bottom:12px;">'
            f'💬 Ask EduBot anything — candidate scores, hiring tips, salary benchmarks, or any general question!</div>',
            unsafe_allow_html=True
        )
    else:
        for msg in st.session_state.chat_history:
            if msg["role"] == "user":
                st.markdown(
                    f'<div style="display:flex;justify-content:flex-end;margin-bottom:10px;">'
                    f'<div style="background:linear-gradient(135deg,#0EA5E9,#0F766E);'
                    f'color:#fff;padding:12px 18px;border-radius:18px 18px 4px 18px;'
                    f'max-width:75%;font-size:.9rem;line-height:1.6;box-shadow:0 2px 12px rgba(14,165,233,.25);">'
                    f'{msg["content"]}</div></div>',
                    unsafe_allow_html=True
                )
            else:
                st.markdown(
                    f'<div style="display:flex;justify-content:flex-start;margin-bottom:10px;">'
                    f'<div style="background:rgba(13,33,64,0.9);border:1px solid rgba(14,165,233,.2);'
                    f'color:{C["cream"]};padding:12px 18px;border-radius:18px 18px 18px 4px;'
                    f'max-width:78%;font-size:.9rem;line-height:1.7;'
                    f'box-shadow:0 2px 12px rgba(0,0,0,.3);">'
                    f'<span style="font-size:.72rem;color:{C["sky"]};font-weight:700;'
                    f'display:block;margin-bottom:5px;">🤖 EduBot</span>'
                    f'{msg["content"]}</div></div>',
                    unsafe_allow_html=True
                )

# ── Chat input row ────────────────────────────────────────────────────────────
chat_cols = st.columns([6, 1])
with chat_cols[0]:
    user_input = st.text_input(
        label="chat_input", label_visibility="collapsed",
        placeholder="Ask about candidates, scores, gaps, recommendations...",
        key="chat_input_field"
    )
with chat_cols[1]:
    send_btn = st.button("Send ➤", use_container_width=True, key="chat_send")

if send_btn and user_input.strip():
    st.session_state.chat_history.append({"role": "user", "content": user_input.strip()})
    with st.spinner("EduBot is thinking..."):
        reply = get_chatbot_response(
            user_message=user_input.strip(),
            chat_history=st.session_state.chat_history[:-1],
            screening_results=st.session_state.chat_results_snapshot,
            job_description=st.session_state.chat_jd_snapshot,
            groq_api_key=groq_key,
            groq_model=groq_model,
            rag_index=st.session_state.get("rag_index"),
        )
    st.session_state.chat_history.append({"role": "assistant", "content": reply})
    st.rerun()

if st.session_state.chat_history:
    if st.button("🗑️ Clear Chat", key="clear_chat"):
        st.session_state.chat_history = []
        st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# FAIRNESS AUDIT — DEI Analysis of Ranked Output
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(divider(), unsafe_allow_html=True)

st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(99,102,241,.12),rgba(139,92,246,.10));
     border:1px solid rgba(99,102,241,.3);border-radius:16px;
     padding:28px 36px;margin-bottom:24px;">
  <div style="display:flex;align-items:center;gap:16px;margin-bottom:10px;">
    <div style="font-size:2rem;">⚖️</div>
    <div>
      <h2 style="font-size:1.5rem;font-weight:800;color:#E8EFF8;margin:0;
        font-family:'DM Serif Display',Georgia,serif;">Fairness & Equity Audit</h2>
      <p style="font-size:.83rem;color:{C['silver']};margin:4px 0 0;">
        Analyse ranked output for patterns across institution tier & degree type · MBA-grade DEI reporting
      </p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if _FAIRNESS_AVAILABLE and st.session_state.get("screening_results"):
    _audit_results = st.session_state["screening_results"]
    _audit_data    = run_fairness_audit(_audit_results)
    st.session_state["fairness_audit"] = _audit_data

    _fa_total  = _audit_data["total_candidates"]
    _fa_qual   = _audit_data["total_qualified"]
    _fa_disps  = _audit_data.get("disparities", [])
    _fa_tier   = _audit_data.get("institution_tier", {})
    _fa_deg    = _audit_data.get("degree_type", {})
    _fa_recs   = _audit_data.get("recommendations", [])

    # Summary banner
    _disp_color = "#EF4444" if _fa_disps else "#10B981"
    _disp_icon  = "⚠️" if _fa_disps else "✅"
    st.markdown(
        f'<div style="background:rgba(0,0,0,.25);border:1.5px solid {_disp_color}55;'
        f'border-radius:12px;padding:16px 22px;margin-bottom:16px;">'
        f'<span style="font-size:1rem;font-weight:800;color:{_disp_color};">'
        f'{_disp_icon} {len(_fa_disps)} statistical disparity flag(s) detected</span>'
        f'<p style="color:#D1DCF0;margin:6px 0 0;font-size:.88rem;">{_audit_data.get("summary","")}</p>'
        f'</div>',
        unsafe_allow_html=True
    )

    _fa_c1, _fa_c2 = st.columns(2)

    with _fa_c1:
        st.markdown(f'<p style="font-size:.95rem;font-weight:700;color:{C["sky"]};margin:0 0 8px;">🏛️ Institution Tier Breakdown</p>', unsafe_allow_html=True)
        for _grp, _gd in _fa_tier.items():
            _bar_color = "#6366F1" if _grp == "Premier" else "#94A3B8"
            _bar_pct   = _gd["pass_rate"]
            st.markdown(
                f'<div style="background:rgba(14,165,233,.07);border:1px solid rgba(14,165,233,.2);'
                f'border-radius:10px;padding:12px 16px;margin-bottom:8px;">'
                f'<div style="display:flex;justify-content:space-between;align-items:center;">'
                f'<span style="font-size:.9rem;font-weight:700;color:#E8EFF8;">{_grp}</span>'
                f'<span style="font-size:.8rem;color:{C["silver"]};">{_gd["count"]} candidates</span>'
                f'</div>'
                f'<div style="display:flex;gap:16px;margin-top:6px;">'
                f'<span style="font-size:.8rem;color:#6EE7B7;">Pass rate: <b>{_gd["pass_rate"]}%</b></span>'
                f'<span style="font-size:.8rem;color:{C["sky"]};">Avg score: <b>{_gd["avg_score"]}%</b></span>'
                + (f'<span style="font-size:.8rem;color:#F59E0B;">Avg rank: <b>{_gd["avg_rank"]}</b></span>' if _gd.get('avg_rank') else '')
                + '</div>'
                f'<div style="background:rgba(255,255,255,.1);border-radius:100px;height:5px;margin-top:8px;">'
                f'<div style="background:{_bar_color};height:5px;border-radius:100px;width:{_bar_pct}%;"></div>'
                f'</div>'
                f'</div>',
                unsafe_allow_html=True
            )

    with _fa_c2:
        st.markdown(f'<p style="font-size:.95rem;font-weight:700;color:{C["sky"]};margin:0 0 8px;">🎓 Degree Type Breakdown</p>', unsafe_allow_html=True)
        for _bucket, _bd in list(_fa_deg.items())[:6]:
            _bc = "#10B981" if _bd["pass_rate"] >= 60 else "#F59E0B" if _bd["pass_rate"] >= 30 else "#EF4444"
            st.markdown(
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'padding:7px 14px;background:rgba(14,165,233,.06);border-radius:8px;margin-bottom:5px;">'
                f'<span style="font-size:.82rem;color:#E8EFF8;font-weight:600;">{_bucket}</span>'
                f'<span style="font-size:.78rem;color:{C["silver"]};">{_bd["count"]} | '
                f'<span style="color:{_bc};font-weight:700;">{_bd["pass_rate"]}% pass</span> | '
                f'{_bd["avg_score"]}% avg</span>'
                f'</div>',
                unsafe_allow_html=True
            )

    # Disparity flags
    if _fa_disps:
        st.markdown(f'<p style="font-size:.9rem;font-weight:700;color:#F59E0B;margin:16px 0 6px;">⚠️ Flagged Disparities</p>', unsafe_allow_html=True)
        for _d in _fa_disps:
            _sev_col = "#EF4444" if _d["severity"] == "High" else "#F59E0B"
            st.markdown(
                f'<div style="background:rgba(239,68,68,.08);border-left:3px solid {_sev_col};'
                f'border-radius:0 8px 8px 0;padding:8px 14px;margin-bottom:6px;">'
                f'<span style="color:{_sev_col};font-weight:700;font-size:.83rem;">{_d["severity"]} severity</span>'
                f'<span style="color:#E8EFF8;font-size:.83rem;"> — {_d["metric"]}: <b>{_d["gap"]}pp gap</b></span>'
                f'</div>',
                unsafe_allow_html=True
            )

    # Recommendations
    if _fa_recs:
        st.markdown(f'<p style="font-size:.9rem;font-weight:700;color:{C["sky"]};margin:14px 0 6px;">💡 Recommendations</p>', unsafe_allow_html=True)
        for _rec in _fa_recs:
            st.markdown(
                f'<div style="padding:6px 0;display:flex;gap:8px;">'
                f'<span style="color:#6366F1;font-size:.85rem;">→</span>'
                f'<span style="font-size:.85rem;color:#D1DCF0;">{_rec}</span>'
                f'</div>',
                unsafe_allow_html=True
            )
elif not _FAIRNESS_AVAILABLE:
    st.info("ℹ️ Fairness audit module not found. Ensure `fairness_audit.py` is in the same directory.")
else:
    st.markdown(
        f'<div style="text-align:center;padding:32px;color:{C["silver"]};font-size:.88rem;'
        f'border:1px dashed rgba(148,163,184,.3);border-radius:12px;">'
        f'⚖️ Run the screening first to generate the fairness audit.</div>',
        unsafe_allow_html=True
    )


# ══════════════════════════════════════════════════════════════════════════════
# EXECUTIVE SUMMARY PDF — One-Click Decision-Support Report
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(divider(), unsafe_allow_html=True)

st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(16,185,129,.12),rgba(6,95,70,.12));
     border:1px solid rgba(16,185,129,.3);border-radius:16px;
     padding:28px 36px;margin-bottom:24px;">
  <div style="display:flex;align-items:center;gap:16px;margin-bottom:10px;">
    <div style="font-size:2rem;">📄</div>
    <div>
      <h2 style="font-size:1.5rem;font-weight:800;color:#E8EFF8;margin:0;
        font-family:'DM Serif Display',Georgia,serif;">Executive Summary PDF</h2>
      <p style="font-size:.83rem;color:{C['silver']};margin:4px 0 0;">
        One-click report a school principal can read · Includes candidate cards, fairness snapshot & filters table
      </p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if _PDF_AVAILABLE and st.session_state.get("screening_results"):
    _pdf_results = st.session_state["screening_results"]
    _pdf_jd      = st.session_state.get("screening_jd", "")
    _pdf_audit   = st.session_state.get("fairness_audit")

    _pdf_c1, _pdf_c2, _pdf_c3 = st.columns([2, 2, 1])
    with _pdf_c1:
        _pdf_school = st.text_input("School Name", value=school_name or "Our School", key="pdf_school")
    with _pdf_c2:
        _pdf_role   = st.text_input("Role / Position", value="Teaching Position", key="pdf_role")
    with _pdf_c3:
        _pdf_hr     = st.text_input("HR Contact Name", value=hr_name or "", key="pdf_hr")

    if st.button("📄 Generate Executive PDF", use_container_width=False, key="gen_pdf_btn"):
        _report_available = _PDF_AVAILABLE
        if not reportlab_available():
            st.error("⚠️ `reportlab` is not installed. Run: `pip install reportlab`")
            _report_available = False

        if _report_available:
            with st.spinner("Building executive PDF report..."):
                _pdf_bytes = generate_executive_pdf(
                    results=_pdf_results,
                    job_description=_pdf_jd,
                    school_name=_pdf_school,
                    role_title=_pdf_role,
                    hr_name=_pdf_hr,
                    audit=_pdf_audit,
                )
            if _pdf_bytes:
                import datetime as _pdt
                _fname = f"EduHire_Report_{_pdt.date.today().strftime('%Y%m%d')}.pdf"
                st.success("✅ PDF ready! Click below to download.")
                st.download_button(
                    label="⬇️ Download Executive Report PDF",
                    data=_pdf_bytes,
                    file_name=_fname,
                    mime="application/pdf",
                    use_container_width=False,
                    key="download_pdf_btn",
                )
            else:
                st.error("PDF generation failed. Check that `reportlab` is installed correctly.")

elif not _PDF_AVAILABLE:
    st.info("ℹ️ `executive_pdf.py` not found. Ensure it is in the same directory as app.py.")
    st.code("pip install reportlab", language="bash")
else:
    st.markdown(
        f'<div style="text-align:center;padding:32px;color:{C["silver"]};font-size:.88rem;'
        f'border:1px dashed rgba(148,163,184,.3);border-radius:12px;">'
        f'📄 Run the screening first to enable PDF generation.</div>',
        unsafe_allow_html=True
    )


# ══════════════════════════════════════════════════════════════════════════════
# JD A/B TESTER — Compare Two Job Descriptions
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(divider(), unsafe_allow_html=True)

st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(245,158,11,.12),rgba(217,119,6,.10));
     border:1px solid rgba(245,158,11,.3);border-radius:16px;
     padding:28px 36px;margin-bottom:24px;">
  <div style="display:flex;align-items:center;gap:16px;margin-bottom:10px;">
    <div style="font-size:2rem;">🔬</div>
    <div>
      <h2 style="font-size:1.5rem;font-weight:800;color:#E8EFF8;margin:0;
        font-family:'DM Serif Display',Georgia,serif;">JD A/B Tester</h2>
      <p style="font-size:.83rem;color:{C['silver']};margin:4px 0 0;">
        Score same resumes against two JDs in parallel · See who surfaces, who drops out, and why
      </p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if _AB_AVAILABLE:
    _ab_c1, _ab_c2 = st.columns(2, gap="large")

    with _ab_c1:
        st.markdown(f'<p style="font-size:.95rem;font-weight:700;color:{C["sky"]};margin:0 0 4px;">📋 JD Version A</p>', unsafe_allow_html=True)
        _ab_label_a = st.text_input("Label for JD A", value="Original JD", key="ab_label_a")
        _ab_jd_a    = st.text_area("Paste JD Version A", height=220, key="ab_jd_a",
                                    placeholder="Paste your original or current job description here…",
                                    label_visibility="collapsed")

    with _ab_c2:
        st.markdown(f'<p style="font-size:.95rem;font-weight:700;color:{C["gold"]};margin:0 0 4px;">📋 JD Version B</p>', unsafe_allow_html=True)
        _ab_label_b = st.text_input("Label for JD B", value="Revised JD", key="ab_label_b")
        _ab_jd_b    = st.text_area("Paste JD Version B", height=220, key="ab_jd_b",
                                    placeholder="Paste your revised or alternative job description here…",
                                    label_visibility="collapsed")

    st.markdown(f'<p style="font-size:.82rem;color:{C["silver"]};margin:6px 0 10px;">Uses the same resumes already uploaded above. Groq scoring applies if your key is set.</p>', unsafe_allow_html=True)

    _ab_run = st.button("🔬 Run A/B Comparison", key="ab_run_btn", use_container_width=False)

    if _ab_run:
        if not _ab_jd_a.strip() or not _ab_jd_b.strip():
            st.error("⚠️ Paste both JD versions before running the A/B test.")
        elif not uploaded_files:
            st.error("⚠️ Upload resumes (in the section above) before running the A/B test.")
        else:
            # Reset file pointers
            for _f in uploaded_files:
                _f.seek(0)
            with st.spinner("Scoring all resumes against both JDs…"):
                _ab_config = {
                    "min_experience_years":    min_exp,
                    "min_similarity_score":    min_score_pct / 100.0,
                    "require_teaching_degree": req_degree,
                    "require_certification":   req_cert,
                }
                _ab_report = run_ab_test(
                    jd_a=_ab_jd_a,
                    jd_b=_ab_jd_b,
                    uploaded_files=uploaded_files,
                    label_a=_ab_label_a or "JD A",
                    label_b=_ab_label_b or "JD B",
                    groq_api_key=groq_key or "",
                    config=_ab_config,
                )
                st.session_state["ab_report"] = _ab_report

    # Display stored A/B report
    _ab_report = st.session_state.get("ab_report")
    if _ab_report:
        _la = _ab_report["label_a"]
        _lb = _ab_report["label_b"]
        _winner = _ab_report.get("winner", "Tie")
        _winner_label = _la if _winner == "A" else (_lb if _winner == "B" else "Tie")
        _winner_color = "#10B981" if _winner != "Tie" else "#94A3B8"

        # KPI strip
        st.markdown("<br>", unsafe_allow_html=True)
        _ab_k1, _ab_k2, _ab_k3, _ab_k4, _ab_k5 = st.columns(5)
        for _col, _val, _lbl, _fg in [
            (_ab_k1, f"{_ab_report['avg_a']:.1f}%",     f"Avg Score ({_la})",   C["sky"]),
            (_ab_k2, f"{_ab_report['avg_b']:.1f}%",     f"Avg Score ({_lb})",   C["gold"]),
            (_ab_k3, str(_ab_report["pass_a"]),          f"Qualified ({_la})",   C["sage"]),
            (_ab_k4, str(_ab_report["pass_b"]),          f"Qualified ({_lb})",   C["sage"]),
            (_ab_k5, _winner_label,                      "Winner",              _winner_color),
        ]:
            _col.markdown(
                f'<div style="background:rgba(14,165,233,.08);border:1px solid {_fg}33;'
                f'border-radius:12px;padding:14px 10px;text-align:center;">'
                f'<div style="font-size:1.6rem;font-weight:900;color:{_fg};">{_val}</div>'
                f'<div style="font-size:.65rem;color:{C["silver"]};margin-top:3px;">{_lbl}</div>'
                f'</div>',
                unsafe_allow_html=True
            )

        st.markdown("<br>", unsafe_allow_html=True)

        # Insights
        _ab_insights = _ab_report.get("insights", [])
        if _ab_insights:
            st.markdown(f'<p style="font-size:.95rem;font-weight:700;color:{C["sky"]};margin:0 0 8px;">💡 Key Insights</p>', unsafe_allow_html=True)
            for _ins in _ab_insights:
                st.markdown(
                    f'<div style="padding:6px 0 6px 4px;display:flex;gap:8px;border-bottom:1px solid rgba(255,255,255,.05);">'
                    f'<span style="color:#F59E0B;font-size:.85rem;">→</span>'
                    f'<span style="font-size:.85rem;color:#D1DCF0;">{_ins}</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )
            st.markdown("<br>", unsafe_allow_html=True)

        # Per-candidate comparison table
        st.markdown(f'<p style="font-size:.95rem;font-weight:700;color:{C["sky"]};margin:0 0 8px;">📊 Candidate Comparison</p>', unsafe_allow_html=True)

        _STATUS_LABELS = {
            "both_pass":  ("✅ Both pass",  "#10B981"),
            "both_fail":  ("❌ Both fail",  "#64748B"),
            "surfaces":   ("🆕 Surfaces",   "#F59E0B"),
            "drops_out":  ("⛔ Drops out",  "#EF4444"),
        }

        for _cand in _ab_report.get("candidates", []):
            _sc_a     = _cand["score_a"]
            _sc_b     = _cand["score_b"]
            _delta    = _cand["delta"]
            _delta_c  = "#10B981" if _delta >= 2 else "#EF4444" if _delta <= -2 else "#94A3B8"
            _delta_s  = f"+{_delta:.1f}pp" if _delta >= 0 else f"{_delta:.1f}pp"
            _st_label, _st_color = _STATUS_LABELS.get(_cand["status_change"], ("—", "#94A3B8"))

            st.markdown(
                f'<div style="background:rgba(13,27,42,.6);border:1px solid rgba(30,58,95,.8);'
                f'border-radius:10px;padding:12px 18px;margin-bottom:8px;">'
                f'<div style="display:flex;justify-content:space-between;align-items:center;">'
                f'<span style="font-size:.92rem;font-weight:700;color:#E8EFF8;">{_cand["name"]}</span>'
                f'<span style="font-size:.78rem;background:{_st_color}22;color:{_st_color};'
                f'padding:3px 10px;border-radius:100px;font-weight:700;">{_st_label}</span>'
                f'</div>'
                f'<div style="display:flex;gap:20px;margin-top:6px;font-size:.82rem;color:{C["silver"]};">'
                f'<span>{_la}: <b style="color:{C["sky"]};">{_sc_a:.1f}%</b></span>'
                f'<span>{_lb}: <b style="color:{C["gold"]};">{_sc_b:.1f}%</b></span>'
                f'<span>Δ: <b style="color:{_delta_c};">{_delta_s}</b></span>'
                f'{"<span>★ Premier</span>" if _cand.get("is_premier") else ""}'
                f'<span>{_cand.get("exp",0)} yr(s)</span>'
                f'</div>'
                f'</div>',
                unsafe_allow_html=True
            )
else:
    st.info("ℹ️ `jd_ab_tester.py` not found. Ensure it is in the same directory as app.py.")


# ══════════════════════════════════════════════════════════════════════════════
# PIPELINE TRACKER — Stage Management, Stall Detection, Cold Monitor, Notes
# ══════════════════════════════════════════════════════════════════════════════

st.markdown(divider(), unsafe_allow_html=True)

st.markdown(f"""
<div style="background:linear-gradient(135deg,rgba(139,92,246,.12),rgba(14,165,233,.10));
     border:1px solid rgba(139,92,246,.3);border-radius:16px;
     padding:28px 36px;margin-bottom:24px;">
  <div style="display:flex;align-items:center;gap:16px;margin-bottom:10px;">
    <div style="font-size:2rem;">🚦</div>
    <div>
      <h2 style="font-size:1.5rem;font-weight:800;color:#E8EFF8;margin:0;
        font-family:'DM Serif Display',Georgia,serif;">Candidate Pipeline Tracker</h2>
      <p style="font-size:.83rem;color:{C['silver']};margin:4px 0 0;">
        Track every candidate through stages · Detect stalls · Monitor cold candidates · Log call notes · Gmail alerts
      </p>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

if "pipeline" not in st.session_state or not st.session_state.get("pipeline"):
    st.markdown(
        f'<div style="text-align:center;padding:32px;color:{C["silver"]};'
        f'font-size:.88rem;border:1px dashed rgba(148,163,184,.3);'
        f'border-radius:12px;margin-bottom:12px;">'
        f'🚦 Run the screening first to populate the pipeline tracker.</div>',
        unsafe_allow_html=True
    )
else:
    _pipeline = st.session_state["pipeline"]
    _all_results = st.session_state.get("screening_results", [])

    # ── Pipeline Summary Bar ──────────────────────────────────────────────────
    summary = get_pipeline_summary(st.session_state)
    st.markdown(
        f'<p style="font-size:1rem;font-weight:700;color:{C["cream"]};margin:0 0 12px;">'
        f'📊 Pipeline Overview</p>',
        unsafe_allow_html=True
    )

    summary_cols = st.columns(len(PIPELINE_STAGES))
    for i, stage in enumerate(PIPELINE_STAGES):
        bg, fg = STAGE_COLORS[stage]
        icon   = STAGE_ICONS[stage]
        count  = summary.get(stage, 0)
        summary_cols[i].markdown(
            f'<div style="background:{bg};border-radius:12px;padding:14px 10px;text-align:center;border:1px solid {fg}44;">'
            f'<div style="font-size:1.6rem;font-weight:900;color:{fg};">{count}</div>'
            f'<div style="font-size:.65rem;color:{fg};text-transform:uppercase;letter-spacing:.8px;margin-top:2px;">{icon} {stage}</div>'
            f'</div>',
            unsafe_allow_html=True
        )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Stall & Cold Alerts ───────────────────────────────────────────────────
    stalled = get_stalled_candidates(st.session_state, stall_days=stall_days)
    cold    = get_cold_candidates(st.session_state, cold_days=cold_days)

    alert_col1, alert_col2 = st.columns(2)

    with alert_col1:
        if stalled:
            st.markdown(
                f'<div style="background:rgba(239,68,68,.1);border:1.5px solid rgba(239,68,68,.4);'
                f'border-radius:12px;padding:16px 20px;">'
                f'<p style="font-size:.92rem;font-weight:800;color:#EF4444;margin:0 0 10px;">'
                f'⚠️ {len(stalled)} Stalled Candidate(s)</p>',
                unsafe_allow_html=True
            )
            for c in stalled:
                st.markdown(
                    f'<div style="display:flex;justify-content:space-between;align-items:center;'
                    f'padding:6px 10px;background:rgba(239,68,68,.08);border-radius:8px;margin-bottom:5px;">'
                    f'<span style="font-size:.82rem;color:#E8EFF8;font-weight:600;">{c["name"]}</span>'
                    f'<span style="font-size:.75rem;color:#EF4444;">{c["_days_stalled"]}d in {c["stage"]}</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                f'<div style="background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.25);'
                f'border-radius:12px;padding:14px 18px;">'
                f'<p style="color:#6EE7B7;font-size:.85rem;margin:0;">✅ No stalled candidates (>{stall_days} days)</p>'
                f'</div>',
                unsafe_allow_html=True
            )

    with alert_col2:
        if cold:
            st.markdown(
                f'<div style="background:rgba(14,165,233,.1);border:1.5px solid rgba(14,165,233,.4);'
                f'border-radius:12px;padding:16px 20px;">'
                f'<p style="font-size:.92rem;font-weight:800;color:#38BDF8;margin:0 0 10px;">'
                f'🧊 {len(cold)} Going Cold</p>',
                unsafe_allow_html=True
            )
            for c in cold:
                st.markdown(
                    f'<div style="display:flex;justify-content:space-between;align-items:center;'
                    f'padding:6px 10px;background:rgba(14,165,233,.08);border-radius:8px;margin-bottom:5px;">'
                    f'<span style="font-size:.82rem;color:#E8EFF8;font-weight:600;">{c["name"]}</span>'
                    f'<span style="font-size:.75rem;color:#38BDF8;">{c["_days_cold"]}d no contact</span>'
                    f'</div>',
                    unsafe_allow_html=True
                )
            st.markdown('</div>', unsafe_allow_html=True)
        else:
            st.markdown(
                f'<div style="background:rgba(16,185,129,.08);border:1px solid rgba(16,185,129,.25);'
                f'border-radius:12px;padding:14px 18px;">'
                f'<p style="color:#6EE7B7;font-size:.85rem;margin:0;">✅ No candidates going cold (>{cold_days} days)</p>'
                f'</div>',
                unsafe_allow_html=True
            )

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Per-Candidate Pipeline Cards ──────────────────────────────────────────
    st.markdown(
        f'<p style="font-size:1rem;font-weight:700;color:{C["cream"]};margin:0 0 4px;">'
        f'👤 Candidate Stage Management</p>'
        f'<p style="font-size:.78rem;color:{C["silver"]};margin:0 0 16px;">'
        f'Change stages, log notes, and get Gmail alerts per candidate.</p>',
        unsafe_allow_html=True
    )

    # Show qualified first, then filtered
    sorted_results = (
        sorted([r for r in _all_results if not r.get("filtered")], key=lambda x: x.get("rank", 999))
        + [r for r in _all_results if r.get("filtered")]
    )

    for cand in sorted_results:
        key     = _cand_key(cand)
        pdata   = _pipeline.get(key, {})
        if not pdata:
            continue

        cur_stage = pdata.get("stage", "Applied")
        bg_c, fg_c = STAGE_COLORS.get(cur_stage, ("#1E3A5F", "#7DD3FC"))
        stage_icon = STAGE_ICONS.get(cur_stage, "📋")
        name       = pdata.get("name", "Unknown")
        score      = pdata.get("score", 0)
        notes_list = pdata.get("notes", [])
        history    = pdata.get("history", [])

        # Days since last contact
        lc_age = days_since(pdata.get("last_contact", ""))
        lc_color = "#EF4444" if lc_age >= stall_days else "#F59E0B" if lc_age >= cold_days else "#10B981"

        with st.expander(
            f"{stage_icon} {name}  ·  {cur_stage}  ·  {score:.0f}%  ·  last contact {lc_age:.0f}d ago",
            expanded=False
        ):
            col_stage, col_meta = st.columns([2, 1])

            with col_stage:
                st.markdown(
                    f'<p style="font-size:.8rem;font-weight:700;color:{C["sky"]};margin:0 0 6px;">📍 Current Stage</p>',
                    unsafe_allow_html=True
                )
                st.markdown(
                    f'<div style="display:inline-block;background:{bg_c};color:{fg_c};'
                    f'padding:6px 18px;border-radius:100px;font-weight:700;font-size:.85rem;'
                    f'border:1px solid {fg_c}55;margin-bottom:12px;">{stage_icon} {cur_stage}</div>',
                    unsafe_allow_html=True
                )

                new_stage = st.selectbox(
                    "Move to stage",
                    PIPELINE_STAGES,
                    index=PIPELINE_STAGES.index(cur_stage),
                    key=f"stage_sel_{key}",
                )

                stage_note = st.text_input(
                    "Optional note for this move",
                    placeholder="e.g. Interview confirmed for Monday",
                    key=f"stage_note_{key}",
                )

                if st.button(f"✅ Update Stage", key=f"stage_btn_{key}", use_container_width=True):
                    if new_stage != cur_stage:
                        gmail_result = set_stage(
                            key, new_stage, st.session_state,
                            note=stage_note,
                            gmail_app_password=resend_api_key,
                            sender_email=sender_email,
                            notify_email=notify_email,
                            school_name=school_name,
                            hr_name=hr_name,
                        )
                        if gmail_result:
                            if gmail_result.get("success"):
                                st.success(f"Stage updated ✅ · {gmail_result['message']}")
                            else:
                                st.warning(f"Stage updated, but email error: {gmail_result['message']}")
                        else:
                            st.success(f"Stage updated → {new_stage}")
                        st.rerun()
                    else:
                        st.info("No change — already in that stage.")

            with col_meta:
                st.markdown(
                    f'<p style="font-size:.8rem;color:{C["silver"]};">'
                    f'📅 Last contact: <span style="color:{lc_color};font-weight:700;">{lc_age:.0f} day(s) ago</span></p>'
                    f'<p style="font-size:.8rem;color:{C["silver"]};">🏅 Score: <b style="color:{C["sky"]};">{score:.0f}%</b></p>'
                    f'<p style="font-size:.8rem;color:{C["silver"]};">📋 Stage changes: <b>{len(history)}</b></p>'
                    f'<p style="font-size:.8rem;color:{C["silver"]};">💬 Notes logged: <b>{len(notes_list)}</b></p>',
                    unsafe_allow_html=True
                )

            # ── Notes Logger ──────────────────────────────────────────────────
            st.markdown(
                "<hr style='border-color:rgba(255,255,255,.1);margin:14px 0;'>",
                unsafe_allow_html=True
            )
            st.markdown(
                f'<p style="font-size:.85rem;font-weight:700;color:{C["sky"]};margin:0 0 6px;">'
                f'📝 Call / Meeting Notes</p>',
                unsafe_allow_html=True
            )

            note_text   = st.text_area(
                "Add note",
                label_visibility="collapsed",
                placeholder="e.g. Spoke with candidate — available from May 12. Strong lesson-planning skills. Follow up by Friday.",
                height=90,
                key=f"note_ta_{key}",
            )
            note_author = st.text_input(
                "Your name (optional)",
                placeholder="Recruiter name",
                key=f"note_author_{key}",
            )

            if st.button("💾 Save Note", key=f"note_btn_{key}"):
                if note_text.strip():
                    add_note(key, note_text.strip(), note_author.strip() or "Recruiter", st.session_state)
                    st.success("Note saved!")
                    st.rerun()
                else:
                    st.warning("Note is empty.")

            # Show existing notes
            if notes_list:
                st.markdown(
                    f'<p style="font-size:.75rem;font-weight:600;color:{C["silver"]};margin:10px 0 6px;">'
                    f'Previous Notes ({len(notes_list)})</p>',
                    unsafe_allow_html=True
                )
                for n in reversed(notes_list):
                    import datetime as _dtm
                    try:
                        ts_dt = _dtm.datetime.fromisoformat(n["ts"])
                        ts_str = ts_dt.strftime("%d %b %Y, %I:%M %p")
                    except Exception:
                        ts_str = n.get("ts", "")
                    st.markdown(
                        f'<div style="background:rgba(14,165,233,.08);border-left:3px solid #0EA5E9;'
                        f'border-radius:0 8px 8px 0;padding:8px 14px;margin-bottom:6px;">'
                        f'<span style="font-size:.7rem;color:#64748B;">{ts_str} · {n.get("author","Recruiter")}</span><br>'
                        f'<span style="font-size:.85rem;color:#E8EFF8;">{n["text"]}</span>'
                        f'</div>',
                        unsafe_allow_html=True
                    )

            # ── Stage History ─────────────────────────────────────────────────
            if len(history) > 1:
                with st.expander("🕐 Stage History", expanded=False):
                    import datetime as _dtm2
                    for h in reversed(history):
                        try:
                            hts = _dtm2.datetime.fromisoformat(h["ts"]).strftime("%d %b %Y, %I:%M %p")
                        except Exception:
                            hts = h.get("ts", "")
                        hbg, hfg = STAGE_COLORS.get(h["stage"], ("#1E3A5F", "#7DD3FC"))
                        hicon = STAGE_ICONS.get(h["stage"], "📋")
                        st.markdown(
                            f'<div style="display:flex;align-items:center;gap:12px;padding:6px 0;border-bottom:1px solid rgba(255,255,255,.05);">'
                            f'<span style="background:{hbg};color:{hfg};font-size:.7rem;padding:2px 10px;border-radius:100px;white-space:nowrap;">{hicon} {h["stage"]}</span>'
                            f'<span style="font-size:.75rem;color:#64748B;">{hts}</span>'
                            f'<span style="font-size:.78rem;color:#A8C5E8;">{h.get("note","")}</span>'
                            f'</div>',
                            unsafe_allow_html=True
                        )




# ─── Footer ───────────────────────────────────────────────────────────────────
st.markdown(
    f'<div style="text-align:center;font-size:.78rem;color:{C["silver"]};'
    f'padding:32px 0 12px;border-top:1px solid {C["border"]};margin-top:48px;">'
    f'EduHire Screener v9 &nbsp;·&nbsp; Groq LLM scoring &nbsp;·&nbsp; '
    f'Fairness Audit &nbsp;·&nbsp; Executive PDF Report &nbsp;·&nbsp; JD A/B Tester &nbsp;·&nbsp; Pipeline Tracker'
    f'</div>',
    unsafe_allow_html=True
)
